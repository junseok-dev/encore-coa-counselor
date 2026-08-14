import csv
import calendar
import base64
import hashlib
import io
import json
import os
import re
import secrets
from time import perf_counter
from urllib.error import HTTPError, URLError
from urllib.parse import quote, urlencode
from urllib.request import Request, urlopen
from uuid import uuid4
from zoneinfo import ZoneInfo
from datetime import date, datetime, time, timedelta
from pathlib import Path

import jwt
from dotenv import dotenv_values
from fastapi import APIRouter, Depends, File, Form, Header, HTTPException, Query, UploadFile
from fastapi.responses import Response
from openai import AsyncOpenAI
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from sqlalchemy import func, inspect as sa_inspect, or_, text
from sqlalchemy.orm import Session

from app.config import ENV_FILE_PATH, get_settings
from app.db.crud import get_all_sessions, get_session_messages
from app.db.database import SessionLocal, get_db
from app.db.models import AdminAuditLog, AdminSecretRecord, AdminUser, AppSetting, BillingCostUploadRecord, BillingDailyCostRecord, CancelRequest, ChatLog, ChatMessage, ChatSession, CustomColumn, CustomTable, DocumentRecord, FaqRecord, OperationsAlert, OperationsAlertHistory, ProcessingLog, PromptConfig, SystemHealthProbe
from app.models.chat import ChatRequest, HistoryMessage
from app.models.session import MessageDetail, SessionDetail, SessionSummary
from app.routers.chat import chat as run_chat_preview
from app.services.admin_service import (
    approve_document,
    create_audit_log,
    full_reindex,
    process_catalog_import,
    process_uploaded_faq_md,
    process_uploaded_md,
    process_uploaded_pdf,
    reject_document,
    restore_document,
    soft_delete_document,
)
from app.services.faq_service import _serialize_faq, seed_faqs, sync_faqs_to_file
from app.services.model_settings import get_active_model, set_active_model
from app.services.prompt_service import PROMPT_DEFAULTS, seed_prompt_configs, serialize_prompt
from app.services.question_category_service import categorize_question_rule, classify_questions_batch
from app.services.storage_service import read_text_from_storage, storage_exists
from app.utils.crypto import ENCRYPTED_PREFIX, decrypt_if_needed, encrypt, maybe_encrypt

router = APIRouter()

ENV_PATH = ENV_FILE_PATH
PROTECTED_PROMPTS = set(PROMPT_DEFAULTS.keys())
NXAVIS_ACCOUNT_ID = "249173798473"
NXAVIS_ACCOUNT_NAME = "엔코아 동작 캠퍼스 5반 30번 학생"

TABLE_DESCRIPTIONS: dict[str, str] = {
    "chat_sessions": "사용자 채팅 세션 목록 — 세션 ID, 생성 시간, 메시지 수 등",
    "chat_messages": "채팅 메시지 내역 — 역할(user/assistant), 소스, 생성 시간 등",
    "chat_logs": "RAG 처리 로그 — 질문, 검색 청크, 답변, LLM 비용 등",
    "documents": "문서 관리 — 업로드·파싱·임베딩·승인 상태 추적",
    "chunks": "문서 청크 — RAG 검색에 사용되는 텍스트 조각",
    "cancel_requests": "취소 요청 내역",
    "operations_alerts": "답변 개선 검토 — 확인 시작·처리 완료 상태 관리",
    "system_health_probes": "시스템 상태 점검 — 데이터베이스 저장 가능 여부 확인",
    "billing_daily_cost_records": "계정·서비스·일자별 실제 원화 사용 금액",
    "billing_cost_upload_records": "월별 n·Xavis 원본 비용 파일 — 마지막 업로드 파일 보관",
    "processing_logs": "문서 처리 로그 — 파싱, 임베딩 등 단계별 처리 결과",
    "prompt_configs": "LLM 프롬프트 설정 — 시스템 프롬프트, 스타일 가이드 등",
    "faqs": "FAQ 데이터 — 질문, 답변, 키워드, 카테고리",
    "admin_audit_logs": "관리자 감사 로그 — 누가, 무엇을, 언제 수행했는지",
    "admin_secret_records": "관리자 보안정보 금고 — 접속 계정과 비밀번호 암호화 저장",
    "custom_tables": "사용자 정의 테이블 메타데이터 (데이터 관리 탭에서 생성한 테이블 목록)",
    "custom_columns": "사용자 정의 테이블 컬럼 정의",
}


def verify_admin(authorization: str = Header(None)) -> str:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="인증이 필요합니다.")
    token = authorization[7:]
    try:
        payload = jwt.decode(token, get_settings().jwt_secret, algorithms=["HS256"])
        return payload["sub"]
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="세션이 만료되었습니다. 다시 로그인해주세요.")
    except Exception:
        raise HTTPException(status_code=401, detail="인증에 실패했습니다.")


def verify_superadmin(current_user: str = Depends(verify_admin)) -> str:
    if current_user != get_settings().admin_email:
        raise HTTPException(status_code=403, detail="최상위 관리자만 접근할 수 있습니다.")
    return current_user


class PasswordChangeRequest(BaseModel):
    new_password: str


class ModelChangeRequest(BaseModel):
    model_name: str


class ReviewRequest(BaseModel):
    note: str | None = None


class FaqItemPayload(BaseModel):
    id: str
    category: str
    question: str
    answer: str
    keywords: list[str] = []
    aliases: list[str] = []
    search_hints: list[str] = []
    source_files: list[str] = []
    direct_answer: bool = False
    top_k: int = 4


class PromptPayload(BaseModel):
    prompt_key: str
    label: str
    content: str


class OperationsAlertUpdate(BaseModel):
    status: str
    note: str | None = None
    test_question: str | None = None
    test_passed: bool | None = None


class OperationsAlertTestRequest(BaseModel):
    question: str


class VaultPasswordRequest(BaseModel):
    password: str


class VaultSecretPayload(BaseModel):
    label: str
    login_url: str | None = None
    account_identifier: str | None = None
    instance_identifier: str | None = None
    username: str | None = None
    password: str | None = None
    note: str | None = None


VAULT_PASSWORD_SETTING_KEY = "admin_security_vault_password_hash"
VAULT_TOKEN_MINUTES = 30
VAULT_DEFAULTS = {
    "nxavis": {
        "label": "n·Xavis 비용 확인",
        "login_url": "https://nxavis.com/layout/usageReport/usageDailyReport",
        "account_identifier": None,
        "instance_identifier": None,
    },
    "aws_console": {
        "label": "AWS 콘솔",
        "login_url": "https://console.aws.amazon.com/",
        "account_identifier": NXAVIS_ACCOUNT_ID,
        "instance_identifier": "i-05ee2bf2ce7f4d0ed",
    },
}
VAULT_ENV_FIELDS = (
    ("OPENAI_API_KEY", "OpenAI API 키", True),
    ("DATABASE_URL", "데이터베이스 연결 주소", True),
    ("AWS_ACCESS_KEY_ID", "AWS 액세스 키 ID", True),
    ("AWS_SECRET_ACCESS_KEY", "AWS 시크릿 액세스 키", True),
    ("AWS_REGION", "AWS 리전", False),
    ("AWS_S3_BUCKET", "S3 버킷", False),
    ("AWS_S3_PREFIX", "S3 경로 접두사", False),
    ("AWS_EC2_INSTANCE_ID", "EC2 인스턴스 ID", False),
    ("GOOGLE_CLIENT_ID", "Google OAuth 클라이언트 ID", False),
    ("CHANNEL_TALK_URL", "채널톡 연결 주소", False),
    ("MODEL_NAME", "기본 LLM 모델", False),
    ("ADMIN_EMAIL", "최상위 관리자 이메일", False),
)


def _hash_vault_password(password: str) -> str:
    iterations = 310_000
    salt = secrets.token_bytes(16)
    digest = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, iterations)
    return "$".join((
        "pbkdf2_sha256",
        str(iterations),
        base64.urlsafe_b64encode(salt).decode("ascii"),
        base64.urlsafe_b64encode(digest).decode("ascii"),
    ))


def _verify_vault_password(password: str, encoded: str) -> bool:
    try:
        algorithm, iterations_text, salt_text, digest_text = encoded.split("$", 3)
        if algorithm != "pbkdf2_sha256":
            return False
        salt = base64.urlsafe_b64decode(salt_text.encode("ascii"))
        expected = base64.urlsafe_b64decode(digest_text.encode("ascii"))
        actual = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt, int(iterations_text))
        return secrets.compare_digest(actual, expected)
    except (TypeError, ValueError):
        return False


def _validate_vault_password(password: str) -> None:
    if len(password) < 8:
        raise HTTPException(status_code=400, detail="보관 비밀번호는 8자 이상이어야 합니다.")
    if not any(not character.isalnum() and not character.isspace() for character in password):
        raise HTTPException(status_code=400, detail="보관 비밀번호에는 특수문자를 1개 이상 포함해야 합니다.")


def _vault_password_setting(db: Session) -> AppSetting | None:
    return db.query(AppSetting).filter(AppSetting.key == VAULT_PASSWORD_SETTING_KEY).first()


def _vault_token_version(db: Session) -> str:
    setting = _vault_password_setting(db)
    if setting is None or not setting.value:
        return ""
    return hashlib.sha256(setting.value.encode("utf-8")).hexdigest()[:24]


def _create_vault_token(current_user: str, db: Session) -> str:
    return jwt.encode(
        {
            "sub": current_user,
            "scope": "admin_security_vault",
            "vault_version": _vault_token_version(db),
            "exp": datetime.now(tz=ZoneInfo("UTC")) + timedelta(minutes=VAULT_TOKEN_MINUTES),
        },
        get_settings().jwt_secret,
        algorithm="HS256",
    )


def _require_vault_token(current_user: str, token: str | None, db: Session) -> None:
    if not token:
        raise HTTPException(status_code=403, detail="보안 정보 잠금을 다시 해제해 주세요.")
    try:
        payload = jwt.decode(token, get_settings().jwt_secret, algorithms=["HS256"])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=403, detail="보안 정보 열람 시간이 만료되었습니다. 다시 잠금 해제해 주세요.")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=403, detail="유효하지 않은 보안 정보 인증입니다.")
    if (
        payload.get("scope") != "admin_security_vault"
        or payload.get("sub") != current_user
        or payload.get("vault_version") != _vault_token_version(db)
    ):
        raise HTTPException(status_code=403, detail="보안 정보 접근 권한이 없습니다.")


def _ensure_vault_defaults(db: Session) -> list[AdminSecretRecord]:
    records: list[AdminSecretRecord] = []
    for secret_key, defaults in VAULT_DEFAULTS.items():
        row = db.query(AdminSecretRecord).filter(AdminSecretRecord.secret_key == secret_key).first()
        if row is None:
            row = AdminSecretRecord(secret_key=secret_key, **defaults)
            db.add(row)
        elif secret_key == "aws_console" and (
            not row.instance_identifier or row.instance_identifier == "i-0123456789abcdef0"
        ):
            row.instance_identifier = defaults["instance_identifier"]
        records.append(row)
    db.flush()
    return records


def _serialize_vault_secret(row: AdminSecretRecord) -> dict:
    return {
        "key": row.secret_key,
        "label": row.label,
        "login_url": row.login_url,
        "account_identifier": row.account_identifier,
        "instance_identifier": row.instance_identifier,
        "username": decrypt_if_needed(row.username_encrypted) or "",
        "password": decrypt_if_needed(row.password_encrypted) or "",
        "note": decrypt_if_needed(row.note_encrypted) or "",
        "updated_by": row.updated_by,
        "updated_at": row.updated_at or row.created_at,
    }


def _vault_environment_items() -> list[dict]:
    file_values = dotenv_values(ENV_PATH) if ENV_PATH.exists() else {}
    items = []
    for key, label, sensitive in VAULT_ENV_FIELDS:
        value = os.getenv(key)
        if value is None:
            value = file_values.get(key)
        text_value = str(value or "")
        items.append({
            "key": key,
            "label": label,
            "value": text_value,
            "configured": bool(text_value),
            "sensitive": sensitive,
        })
    return items


def _serialize_document(record: DocumentRecord) -> dict:
    return {
        "id": record.id,
        "logical_name": record.logical_name,
        "version": record.version,
        "original_filename": decrypt_if_needed(record.original_filename) or "",
        "status": record.status,
        "parser_type": record.parser_type,
        "is_active": record.is_active,
        "is_deleted": getattr(record, "is_deleted", False),
        "review_note": decrypt_if_needed(getattr(record, "review_note", None)),
        "approved_at": getattr(record, "approved_at", None),
        "rejected_at": getattr(record, "rejected_at", None),
        "deleted_at": getattr(record, "deleted_at", None),
        "error_message": decrypt_if_needed(record.error_message),
        "created_at": record.created_at,
        "updated_at": record.updated_at,
        "has_md": storage_exists(record.md_path),
        "has_json": storage_exists(record.json_path),
        "has_pdf": storage_exists(record.pdf_path) or bool(record.storage_key),
    }


def _read_optional_text(path_value: str | None) -> str | None:
    return read_text_from_storage(path_value)


def _serialize_processing_log(row: ProcessingLog) -> dict:
    return {
        "id": row.id,
        "document_id": row.document_id,
        "log_type": row.log_type,
        "status": row.status,
        "message": decrypt_if_needed(row.message) or "",
        "detail": decrypt_if_needed(row.detail),
        "created_at": row.created_at,
    }


def _serialize_chat_log(row: ChatLog) -> dict:
    retrieval_chunks = decrypt_if_needed(row.retrieval_chunks) or "[]"
    return {
        "id": row.id,
        "session_id": row.session_id,
        "question": decrypt_if_needed(row.question) or "",
        "retrieval_chunks": json.loads(retrieval_chunks or "[]"),
        "answer": decrypt_if_needed(row.answer) or "",
        "source": row.source,
        "error": decrypt_if_needed(row.error),
        "processing_status": row.processing_status,
        "question_category": row.question_category,
        "question_category_label": row.question_category_label,
        "question_category_source": row.question_category_source,
        "embedding_cost": row.embedding_cost,
        "llm_cost": row.llm_cost,
        "created_at": row.created_at,
    }


def _serialize_audit_log(row: AdminAuditLog) -> dict:
    return {
        "id": row.id,
        "actor": row.actor,
        "action": row.action,
        "target_type": row.target_type,
        "target_id": row.target_id,
        "detail": decrypt_if_needed(row.detail),
        "created_at": row.created_at,
    }


def _serialize_billing_cost_upload(row: BillingCostUploadRecord | None) -> dict | None:
    if row is None:
        return None
    return {
        "id": row.id,
        "billing_month": row.billing_month,
        "account_id": row.account_id,
        "account_name": row.account_name,
        "filename": row.original_filename,
        "content_type": row.content_type,
        "size_bytes": row.size_bytes,
        "imported_rows": row.imported_rows,
        "uploaded_by": row.uploaded_by,
        "uploaded_at": row.updated_at or row.created_at,
    }


def _normalize_cost_header(value: object) -> str:
    return re.sub(r"[\s_()\-]", "", str(value or "").strip().lower())


def _parse_cost_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    normalized = str(value or "").strip().replace(".", "-").replace("/", "-")
    return datetime.strptime(normalized, "%Y-%m-%d").date()


def _parse_krw_amount(value: object) -> int:
    if isinstance(value, (int, float)):
        return round(value)
    normalized = re.sub(r"[^0-9.\-]", "", str(value or ""))
    return round(float(normalized or "0"))


def _percent_change(current: int, previous: int) -> float | None:
    if previous == 0:
        return 100.0 if current > 0 else 0.0
    return round(((current - previous) / previous) * 100, 1)


def _is_refund_request(value: str | None) -> bool:
    normalized = "".join((value or "").lower().split())
    return any(token in normalized for token in ("환불", "환급", "돈돌려", "결제취소", "수강료돌려"))


def _is_homepage_request(value: str | None) -> bool:
    normalized = "".join((value or "").lower().split())
    return any(token in normalized for token in (
        "홈페이지", "홈피", "공식사이트", "사이트주소", "홈페이지주소", "홈페이지링크", "공식링크",
    ))


def _handoff_reason(
    question: str,
    status: str,
    is_cancel: bool,
    answer: str = "",
    session_context: str = "",
) -> tuple[str, str]:
    normalized = "".join(f"{session_context} {question} {answer}".lower().split())
    if is_cancel:
        return "cancel", "취소 요청"
    if any(token in normalized for token in ("취업", "합격", "채용", "책임", "보장")):
        return "employment", "취업·채용 상담"
    if any(token in normalized for token in ("환불", "결제", "수강료", "비용", "입금")):
        return "payment", "환불·결제 문의"
    if any(token in normalized for token in ("신청", "지원", "접수", "등록", "선발", "합격발표")):
        return "application", "신청·등록 문의"
    if any(token in normalized for token in ("출결", "출석", "결석", "지각", "조퇴", "휴가", "일정", "개강", "수업시간")):
        return "schedule", "일정·출결 문의"
    if any(token in normalized for token in ("커리큘럼", "과정", "강의", "수업", "교육내용", "난이도", "프로젝트")):
        return "course", "과정·커리큘럼 상담"
    if any(token in normalized for token in ("자격", "전공", "비전공", "나이", "국비", "내일배움", "수강조건")):
        return "eligibility", "지원 자격 상담"
    if any(token in normalized for token in ("캠퍼스", "위치", "주소", "주차", "시설", "강의장")):
        return "campus", "캠퍼스·시설 문의"
    if any(token in normalized for token in ("서류", "증명서", "수료증", "행정", "제출", "발급")):
        return "document", "서류·행정 문의"
    if any(token in normalized for token in ("오류", "에러", "로그인", "접속", "저장", "데이터", "사이트", "시스템")):
        return "technical", "기술·시스템 문제"
    if any(token in normalized for token in ("상담", "상담원", "매니저", "사람", "전화", "연락")):
        return "direct", "상담원 직접 요청"
    category = categorize_question_rule(normalized)
    if category.key != "general":
        return f"question_{category.key}", f"{category.label} 상담"
    if status == "handoff_offer":
        return "bot_offer", "답변 한계로 상담 권유"
    return "review_needed", "내용 확인 필요"


def _operations_summary(
    sessions: list[ChatSession],
    logs: list[ChatLog],
    cancels: list[CancelRequest],
) -> dict:
    handoffs = [row for row in logs if row.processing_status == "handoff" or row.source == "handoff"]
    safety = [row for row in logs if row.source == "guardrail"]
    failed = [row for row in logs if row.processing_status == "failed" or bool(row.error)]
    refund_count = sum(1 for row in cancels if _is_refund_request(decrypt_if_needed(row.message)))
    homepage_request_count = sum(
        1 for row in logs if _is_homepage_request(decrypt_if_needed(row.question))
    )
    return {
        "visitors": len(sessions),
        "chats": len(logs),
        "handoffs": len(handoffs),
        "cancels": len(cancels) - refund_count,
        "refunds": refund_count,
        "homepage_requests": homepage_request_count,
        "safety": len(safety),
        "failed": len(failed),
        "avg_questions_per_session": round(len(logs) / len(sessions), 1) if sessions else 0,
    }


def _question_category(question: str) -> tuple[str, str]:
    category = categorize_question_rule(question)
    return category.key, category.label


def _shift_month(month_start: date, offset: int) -> date:
    month_index = month_start.year * 12 + (month_start.month - 1) + offset
    return date(month_index // 12, month_index % 12 + 1, 1)


def _analysis_datetime(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value
    return value.astimezone(ZoneInfo("Asia/Seoul"))


def _peak_item(items: list[dict], key: str, label_key: str) -> dict | None:
    if not items or max(item[key] for item in items) == 0:
        return None
    peak = max(items, key=lambda item: item[key])
    return {"label": peak[label_key], "count": peak[key]}


def _ec2_health_check() -> dict:
    settings = get_settings()
    checked_at = datetime.now()
    if not settings.aws_ec2_instance_id:
        return {
            "key": "ec2",
            "label": "EC2",
            "status": "not_configured",
            "message": "인스턴스 ID를 설정하면 AWS 상태 검사를 표시합니다.",
            "latency_ms": None,
            "checked_at": checked_at,
            "details": {
                "instance_id": None,
                "cause": "AWS_EC2_INSTANCE_ID가 백엔드 환경 변수에 설정되지 않았습니다.",
                "action": "보안 정보 또는 백엔드 .env에 실제 EC2 인스턴스 ID를 설정하고 백엔드를 재시작하세요.",
            },
        }

    started = perf_counter()
    try:
        import boto3
        from botocore.config import Config

        session_kwargs = {"region_name": settings.aws_region}
        if settings.aws_access_key_id and settings.aws_secret_access_key:
            session_kwargs.update({
                "aws_access_key_id": settings.aws_access_key_id,
                "aws_secret_access_key": settings.aws_secret_access_key,
            })
        client = boto3.session.Session(**session_kwargs).client(
            "ec2",
            config=Config(connect_timeout=2, read_timeout=3, retries={"max_attempts": 1}),
        )
        response = client.describe_instance_status(
            InstanceIds=[settings.aws_ec2_instance_id],
            IncludeAllInstances=True,
        )
        statuses = response.get("InstanceStatuses", [])
        if not statuses:
            raise RuntimeError("instance status not found")
        instance = statuses[0]
        state = instance.get("InstanceState", {}).get("Name", "unknown")
        system_status = instance.get("SystemStatus", {}).get("Status", "insufficient-data")
        instance_status = instance.get("InstanceStatus", {}).get("Status", "insufficient-data")
        attached_ebs_status = instance.get("AttachedEbsStatus", {}).get("Status", "not-applicable")
        events = instance.get("Events", [])
        is_healthy = (
            state == "running"
            and system_status == "ok"
            and instance_status == "ok"
            and attached_ebs_status in {"ok", "not-applicable"}
            and not events
        )
        return {
            "key": "ec2",
            "label": "EC2",
            "status": "healthy" if is_healthy else "critical",
            "message": "인스턴스와 AWS 기반 시설 상태가 정상입니다." if is_healthy else "EC2 상태 검사 또는 예약 작업을 확인하세요.",
            "latency_ms": round((perf_counter() - started) * 1000),
            "checked_at": checked_at,
            "details": {
                "instance_id": settings.aws_ec2_instance_id,
                "state": state,
                "system_status": system_status,
                "instance_status": instance_status,
                "attached_ebs_status": attached_ebs_status,
                "scheduled_events": len(events),
            },
        }
    except Exception as exc:
        error_code = ""
        try:
            error_code = str(exc.response.get("Error", {}).get("Code", ""))
        except (AttributeError, TypeError):
            pass
        access_denied = error_code in {"AccessDenied", "AccessDeniedException", "UnauthorizedOperation"}
        return {
            "key": "ec2",
            "label": "EC2",
            "status": "unknown",
            "message": (
                "EC2 장애가 아니라 IAM 상태 조회 권한이 없어 확인할 수 없습니다. "
                "AWS 관리자에게 ec2:DescribeInstanceStatus 읽기 권한을 요청하세요."
                if access_denied else
                "AWS 상태를 조회하지 못했습니다. IAM 권한과 인스턴스 설정을 확인하세요."
            ),
            "latency_ms": round((perf_counter() - started) * 1000),
            "checked_at": checked_at,
            "details": {
                "instance_id": settings.aws_ec2_instance_id,
                "region": settings.aws_region,
                "error_type": type(exc).__name__,
                "error_code": error_code or None,
                "cause": (
                    "현재 AWS 자격 증명에 EC2 상태 조회 권한이 없습니다. EC2 장애를 의미하지는 않습니다."
                    if access_denied else
                    "AWS 자격 증명, 리전, 네트워크 또는 인스턴스 설정 때문에 상태 조회 요청이 실패했습니다."
                ),
                "action": (
                    "AWS 관리자에게 ec2:DescribeInstanceStatus 허용을 요청하세요. SCP의 명시적 거부가 있으면 IAM 권한만 추가해도 해결되지 않습니다."
                    if access_denied else
                    "AWS 자격 증명과 리전, 인스턴스 ID를 확인한 뒤 다시 점검하세요."
                ),
                "required_permission": "ec2:DescribeInstanceStatus",
            },
        }


def _crypt_value(value: str | None, should_encrypt: bool) -> str | None:
    if value is None:
        return None
    if not value:
        return value
    if should_encrypt:
        return maybe_encrypt(value)
    if value.startswith(ENCRYPTED_PREFIX):
        return decrypt_if_needed(value)
    return value


def _upsert_faq_row(db: Session, payload: FaqItemPayload) -> FaqRecord:
    row = db.query(FaqRecord).filter(FaqRecord.faq_key == payload.id).first()
    enc = get_settings().encrypt_faq
    values = {
        "category": _crypt_value(payload.category, enc),
        "question": _crypt_value(payload.question, enc),
        "answer": _crypt_value(payload.answer, enc),
        "keywords_json": _crypt_value(json.dumps(payload.keywords, ensure_ascii=False), enc),
        "aliases_json": _crypt_value(json.dumps(payload.aliases, ensure_ascii=False), enc),
        "search_hints_json": _crypt_value(json.dumps(payload.search_hints, ensure_ascii=False), enc),
        "source_files_json": _crypt_value(json.dumps(payload.source_files, ensure_ascii=False), enc),
        "direct_answer": payload.direct_answer,
        "top_k": payload.top_k,
        "is_active": True,
    }
    if row:
        for key, value in values.items():
            setattr(row, key, value)
    else:
        row = FaqRecord(faq_key=payload.id, **values)
        db.add(row)
    db.commit()
    db.refresh(row)
    create_audit_log(db, "faq_saved", "faq", payload.id, payload.question)
    return row


def _build_workbook(rows: list[dict]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "chat_logs"
    sheet.append(["session_id", "question", "answer", "source", "processing_status", "embedding_cost", "llm_cost", "created_at"])
    for row in rows:
        sheet.append(
            [
                row["session_id"],
                row["question"],
                row["answer"],
                row["source"],
                row["processing_status"],
                row["embedding_cost"],
                row["llm_cost"],
                row["created_at"].isoformat() if row["created_at"] else "",
            ]
        )
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def _filter_chat_logs(db: Session, start_date: date | None = None, end_date: date | None = None, session_id: str | None = None, limit: int | None = 500) -> list[ChatLog]:
    query = db.query(ChatLog)
    if start_date:
        query = query.filter(ChatLog.created_at >= datetime.combine(start_date, time.min))
    if end_date:
        query = query.filter(ChatLog.created_at <= datetime.combine(end_date, time.max))
    if session_id:
        query = query.filter(ChatLog.session_id == session_id)
    query = query.order_by(ChatLog.created_at.desc())
    if limit is not None:  # limit=None이면 전량 (엑셀 내보내기용)
        query = query.limit(limit)
    return query.all()


@router.get("/sessions", response_model=list[SessionSummary])
def list_sessions(skip: int = 0, limit: int = 50, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    sessions = get_all_sessions(db, skip=skip, limit=limit)
    result = []
    for session in sessions:
        summary = SessionSummary.model_validate(session)
        summary.user_name = decrypt_if_needed(session.encrypted_user_name) if session.encrypted_user_name else None
        result.append(summary)
    return result


@router.get("/sessions/{session_id}", response_model=SessionDetail)
def get_session_detail(session_id: str, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    session = db.query(ChatSession).filter(ChatSession.id == session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="세션을 찾을 수 없습니다.")
    messages = get_session_messages(db, session_id)
    summary = SessionSummary.model_validate(session)
    summary.user_name = decrypt_if_needed(session.encrypted_user_name) if session.encrypted_user_name else None
    decrypted_messages = []
    for message in messages:
        detail = MessageDetail.model_validate(message)
        detail.content = decrypt_if_needed(message.content) or ""
        decrypted_messages.append(detail)
    return SessionDetail(session=summary, messages=decrypted_messages)


@router.post("/upload-md")
async def upload_md(file: UploadFile = File(...), title: str = Form(None), category: str = Form(None), db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    if not file.filename or not file.filename.lower().endswith(".md"):
        raise HTTPException(status_code=400, detail="MD 파일만 업로드할 수 있습니다.")
    record = await process_uploaded_md(db, file.filename, await file.read(), title=title, category=category)
    return {"message": "MD 업로드 후 검토 대기 상태로 저장했습니다.", "document": _serialize_document(record)}


@router.post("/upload-faq-md")
async def upload_faq_md(file: UploadFile = File(...), category: str = Form(None), db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    if not file.filename or not file.filename.lower().endswith(".md"):
        raise HTTPException(status_code=400, detail="MD 파일만 업로드할 수 있습니다.")
    record, faq_items = await process_uploaded_faq_md(db, file.filename, await file.read(), category=category)
    return {
        "message": "FAQ 변환 결과를 생성했고, 아직 운영 반영 전입니다.",
        "document": _serialize_document(record),
        "faqs": faq_items,
    }


@router.post("/import-catalog")
async def import_catalog(catalog: UploadFile = File(...), files: list[UploadFile] = File(...), db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    if not catalog.filename or not catalog.filename.lower().endswith(".json"):
        raise HTTPException(status_code=400, detail="catalog는 JSON 파일이어야 합니다.")
    catalog_data = json.loads(await catalog.read())
    md_files = {f.filename: await f.read() for f in files if f.filename and f.filename.lower().endswith(".md")}
    records = await process_catalog_import(db, catalog_data, md_files)
    return {"message": f"{len(records)}개 문서를 검토 대기 상태로 가져왔습니다.", "documents": [_serialize_document(r) for r in records]}


@router.post("/upload-pdf")
async def upload_pdf(file: UploadFile = File(...), db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    if not file.filename or not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="PDF 파일만 업로드할 수 있습니다.")
    record = await process_uploaded_pdf(db, file.filename, await file.read())
    return {"message": "PDF 업로드와 MD 변환이 완료되었고, 현재 검토 대기 상태입니다.", "document": _serialize_document(record)}


@router.get("/documents")
def list_documents(
    parser_type: str | None = Query(default=None),
    include_deleted: bool = Query(default=False),
    status: str | None = Query(default=None),
    db: Session = Depends(get_db),
    _: None = Depends(verify_admin),
):
    query = db.query(DocumentRecord).order_by(DocumentRecord.created_at.desc())
    if parser_type:
        query = query.filter(DocumentRecord.parser_type == parser_type)
    if not include_deleted:
        query = query.filter(DocumentRecord.is_deleted.is_(False))
    if status:
        query = query.filter(DocumentRecord.status == status)
    return {"documents": [_serialize_document(row) for row in query.all()]}


@router.get("/documents/{document_id}")
def get_document_detail(document_id: int, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    record = db.query(DocumentRecord).filter(DocumentRecord.id == document_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="문서를 찾을 수 없습니다.")
    return {"document": _serialize_document(record), "md_content": _read_optional_text(record.md_path), "json_content": _read_optional_text(record.json_path)}


@router.post("/documents/{document_id}/approve")
def approve_document_route(document_id: int, body: ReviewRequest, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    record = db.query(DocumentRecord).filter(DocumentRecord.id == document_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="문서를 찾을 수 없습니다.")
    updated = approve_document(db, record, body.note)
    return {"message": "문서를 승인해 운영 데이터에 반영했습니다.", "document": _serialize_document(updated)}


@router.post("/documents/{document_id}/reject")
def reject_document_route(document_id: int, body: ReviewRequest, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    record = db.query(DocumentRecord).filter(DocumentRecord.id == document_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="문서를 찾을 수 없습니다.")
    updated = reject_document(db, record, body.note)
    return {"message": "문서를 반려했습니다.", "document": _serialize_document(updated)}


@router.post("/documents/{document_id}/restore")
def restore_document_route(document_id: int, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    record = db.query(DocumentRecord).filter(DocumentRecord.id == document_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="문서를 찾을 수 없습니다.")
    updated = restore_document(db, record)
    return {"message": "문서를 복구해 다시 검토 대기 상태로 돌렸습니다.", "document": _serialize_document(updated)}


@router.delete("/documents/{document_id}")
def delete_document(document_id: int, note: str | None = Query(default=None), db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    record = db.query(DocumentRecord).filter(DocumentRecord.id == document_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="문서를 찾을 수 없습니다.")
    updated = soft_delete_document(db, record, note)
    return {"message": "문서를 삭제 처리했습니다.", "document": _serialize_document(updated)}


@router.post("/documents/{document_id}/retry")
def retry_document(document_id: int, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    record = db.query(DocumentRecord).filter(DocumentRecord.id == document_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="문서를 찾을 수 없습니다.")
    if record.status != "failed":
        return {"message": "현재 문서는 재처리 대상이 아닙니다."}
    return {"message": "재처리는 같은 파일을 다시 업로드하는 방식으로 진행합니다."}


@router.post("/reindex")
def reindex(db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    full_reindex(db)
    create_audit_log(db, "reindex", "system", "global", "full_rebuild")
    return {"message": "전체 인덱스를 다시 생성했습니다.", "strategy": "full_rebuild"}


@router.get("/faqs")
def get_faqs(db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    seed_faqs(db)
    rows = db.query(FaqRecord).filter(FaqRecord.is_active.is_(True)).order_by(FaqRecord.id.asc()).all()
    return {"faqs": [_serialize_faq(row) for row in rows]}


@router.post("/faqs")
def create_faq(body: FaqItemPayload, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    row = _upsert_faq_row(db, body)
    sync_faqs_to_file(db)
    full_reindex(db)
    return {"message": "FAQ를 추가했습니다.", "faq": _serialize_faq(row)}


@router.put("/faqs/{faq_key}")
def update_faq(faq_key: str, body: FaqItemPayload, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    if faq_key != body.id:
        raise HTTPException(status_code=400, detail="FAQ 키가 일치하지 않습니다.")
    row = _upsert_faq_row(db, body)
    sync_faqs_to_file(db)
    full_reindex(db)
    return {"message": "FAQ를 수정했습니다.", "faq": _serialize_faq(row)}


@router.delete("/faqs/{faq_key}")
def delete_faq(faq_key: str, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    row = db.query(FaqRecord).filter(FaqRecord.faq_key == faq_key).first()
    if not row:
        raise HTTPException(status_code=404, detail="FAQ를 찾을 수 없습니다.")
    row.is_active = False
    db.commit()
    sync_faqs_to_file(db)
    full_reindex(db)
    create_audit_log(db, "faq_deleted", "faq", faq_key)
    return {"message": "FAQ를 삭제했습니다."}


@router.get("/prompts")
def get_prompts(db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    seed_prompt_configs(db)
    prompts = db.query(PromptConfig).order_by(PromptConfig.id.asc()).all()
    return {"prompts": [serialize_prompt(row) for row in prompts]}


@router.post("/prompts")
def create_prompt(body: PromptPayload, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    existing = db.query(PromptConfig).filter(PromptConfig.prompt_key == body.prompt_key).first()
    if existing:
        raise HTTPException(status_code=409, detail="같은 키의 프롬프트가 이미 있습니다.")
    enc = get_settings().encrypt_prompt
    stored_content = encrypt(body.content) if enc else body.content
    row = PromptConfig(prompt_key=body.prompt_key, label=body.label, content=stored_content)
    db.add(row)
    db.commit()
    db.refresh(row)
    create_audit_log(db, "prompt_created", "prompt", body.prompt_key, body.label)
    return {"message": "프롬프트를 추가했습니다.", "prompt": serialize_prompt(row)}


@router.put("/prompts/{prompt_key}")
def update_prompt(prompt_key: str, body: PromptPayload, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    if prompt_key != body.prompt_key:
        raise HTTPException(status_code=400, detail="프롬프트 키가 일치하지 않습니다.")
    row = db.query(PromptConfig).filter(PromptConfig.prompt_key == prompt_key).first()
    if not row:
        raise HTTPException(status_code=404, detail="프롬프트를 찾을 수 없습니다.")
    enc = get_settings().encrypt_prompt
    row.label = body.label
    row.content = encrypt(body.content) if enc else body.content
    db.commit()
    db.refresh(row)
    create_audit_log(db, "prompt_updated", "prompt", body.prompt_key, body.label)
    return {"message": "프롬프트를 수정했습니다.", "prompt": serialize_prompt(row)}


@router.delete("/prompts/{prompt_key}")
def delete_prompt(prompt_key: str, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    if prompt_key in PROTECTED_PROMPTS:
        raise HTTPException(status_code=400, detail="기본 시스템 프롬프트는 삭제할 수 없습니다.")
    row = db.query(PromptConfig).filter(PromptConfig.prompt_key == prompt_key).first()
    if not row:
        raise HTTPException(status_code=404, detail="프롬프트를 찾을 수 없습니다.")
    db.delete(row)
    db.commit()
    create_audit_log(db, "prompt_deleted", "prompt", prompt_key)
    return {"message": "프롬프트를 삭제했습니다."}


@router.get("/logs")
def get_logs(limit: int = 100, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    processing_logs = db.query(ProcessingLog).order_by(ProcessingLog.created_at.desc()).limit(limit).all()
    chat_logs = db.query(ChatLog).order_by(ChatLog.created_at.desc()).limit(limit).all()
    audit_logs = db.query(AdminAuditLog).order_by(AdminAuditLog.created_at.desc()).limit(limit).all()
    return {
        "processing_logs": [_serialize_processing_log(row) for row in processing_logs],
        "chat_logs": [_serialize_chat_log(row) for row in chat_logs],
        "audit_logs": [_serialize_audit_log(row) for row in audit_logs],
    }


@router.get("/audit-logs")
def get_audit_logs(limit: int = 100, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    rows = db.query(AdminAuditLog).order_by(AdminAuditLog.created_at.desc()).limit(limit).all()
    return {"audit_logs": [_serialize_audit_log(row) for row in rows]}


@router.get("/chat-logs")
def list_chat_logs(start_date: date | None = None, end_date: date | None = None, session_id: str | None = None, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    rows = _filter_chat_logs(db, start_date=start_date, end_date=end_date, session_id=session_id)
    return {"chat_logs": [_serialize_chat_log(row) for row in rows]}


@router.get("/chat-logs/export")
def export_chat_logs(start_date: date | None = None, end_date: date | None = None, session_id: str | None = None, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    rows = [_serialize_chat_log(row) for row in _filter_chat_logs(db, start_date=start_date, end_date=end_date, session_id=session_id, limit=None)]
    payload = _build_workbook(rows)
    filename = f"chat_logs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=payload,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/operations/analytics")
def get_operations_analytics(
    selected_year: int | None = Query(None, ge=2026, le=2100),
    selected_month: str | None = Query(None, pattern=r"^\d{4}-(0[1-9]|1[0-2])$"),
    db: Session = Depends(get_db),
    _: None = Depends(verify_admin),
):
    today = datetime.now().date()
    current_month = today.replace(day=1)

    date_ranges = [
        db.query(func.min(model.created_at), func.max(model.created_at)).one()
        for model in (ChatSession, ChatLog, CancelRequest)
    ]
    latest_values = [value for _, maximum in date_ranges if (value := maximum) is not None]
    latest_data_month = (
        max(_analysis_datetime(value).date() for value in latest_values).replace(day=1)
        if latest_values else current_month
    )
    first_available_year = 2026
    last_available_year = max(today.year, first_available_year)
    available_month_starts = [
        date(year, month, 1)
        for year in range(first_available_year, last_available_year + 1)
        for month in range(1, 13)
    ]
    available_months = [month.strftime("%Y-%m") for month in available_month_starts]

    if selected_month:
        selected_start = datetime.strptime(selected_month, "%Y-%m").date()
        if selected_year is not None and selected_start.year != selected_year:
            raise HTTPException(status_code=400, detail="선택한 연도와 월이 일치하지 않습니다.")
        month_starts = [selected_start]
        analysis_start = datetime.combine(selected_start, time.min)
        analysis_end = datetime.combine(_shift_month(selected_start, 1), time.min)
    elif selected_year is not None:
        selected_start = date(selected_year, 1, 1)
        month_starts = [_shift_month(selected_start, offset) for offset in range(12)]
        analysis_start = datetime.combine(selected_start, time.min)
        analysis_end = datetime.combine(date(selected_year + 1, 1, 1), time.min)
    else:
        first_month = date(first_available_year, 1, 1)
        last_month = max(current_month, latest_data_month)
        month_count = (last_month.year - first_month.year) * 12 + last_month.month
        month_starts = [_shift_month(first_month, offset) for offset in range(month_count)]
        analysis_start = datetime.combine(first_month, time.min)
        analysis_end = datetime.combine(_shift_month(last_month, 1), time.min)

    # 시간대·질문·응답·상담 분석도 선택한 월 또는 전체 기간과 동일하게 맞춥니다.
    query_start = analysis_start - timedelta(days=1)
    query_end = analysis_end + timedelta(days=1)
    sessions = db.query(ChatSession).filter(
        ChatSession.created_at >= query_start,
        ChatSession.created_at < query_end,
    ).all()
    logs = db.query(ChatLog).filter(
        ChatLog.created_at >= query_start,
        ChatLog.created_at < query_end,
    ).all()
    cancels = db.query(CancelRequest).filter(
        CancelRequest.created_at >= query_start,
        CancelRequest.created_at < query_end,
    ).all()

    def in_analysis_period(value: datetime | None) -> bool:
        if not value:
            return False
        local_value = _analysis_datetime(value).replace(tzinfo=None)
        return analysis_start <= local_value < analysis_end

    analysis_sessions = [row for row in sessions if in_analysis_period(row.created_at)]
    analysis_logs = [row for row in logs if in_analysis_period(row.created_at)]
    analysis_cancels = [row for row in cancels if in_analysis_period(row.created_at)]

    monthly_map = {
        month_start.strftime("%Y-%m"): {
            "month": month_start.strftime("%Y-%m"),
            "visitors": 0,
            "chats": 0,
            "handoffs": 0,
            "cancels": 0,
            "refunds": 0,
            "homepage_requests": 0,
            "safety": 0,
            "failed": 0,
        }
        for month_start in month_starts
    }
    hourly_map = {
        hour: {"hour": hour, "label": f"{hour:02d}시", "visitors": 0, "chats": 0}
        for hour in range(24)
    }
    daily_map: dict[str, dict] = {}
    if selected_month:
        _, day_count = calendar.monthrange(selected_start.year, selected_start.month)
        daily_map = {
            date(selected_start.year, selected_start.month, day).isoformat(): {
                "date": date(selected_start.year, selected_start.month, day).isoformat(),
                "visitors": 0,
                "chats": 0,
                "handoffs": 0,
                "cancels": 0,
                "refunds": 0,
                "homepage_requests": 0,
                "safety": 0,
                "failed": 0,
            }
            for day in range(1, day_count + 1)
        }

    for row in analysis_sessions:
        if not row.created_at:
            continue
        local_created_at = _analysis_datetime(row.created_at)
        month_key = local_created_at.strftime("%Y-%m")
        if month_key in monthly_map:
            monthly_map[month_key]["visitors"] += 1
        hourly_map[local_created_at.hour]["visitors"] += 1
        day_key = local_created_at.date().isoformat()
        if day_key in daily_map:
            daily_map[day_key]["visitors"] += 1

    for row in analysis_logs:
        if not row.created_at:
            continue
        local_created_at = _analysis_datetime(row.created_at)
        month_key = local_created_at.strftime("%Y-%m")
        if month_key in monthly_map:
            monthly_map[month_key]["chats"] += 1
            if row.processing_status == "handoff" or row.source == "handoff":
                monthly_map[month_key]["handoffs"] += 1
            if row.source == "guardrail":
                monthly_map[month_key]["safety"] += 1
            if row.processing_status == "failed" or bool(decrypt_if_needed(row.error)):
                monthly_map[month_key]["failed"] += 1
            if _is_homepage_request(decrypt_if_needed(row.question)):
                monthly_map[month_key]["homepage_requests"] += 1
        hourly_map[local_created_at.hour]["chats"] += 1
        day_key = local_created_at.date().isoformat()
        if day_key in daily_map:
            daily_map[day_key]["chats"] += 1
            if row.processing_status == "handoff" or row.source == "handoff":
                daily_map[day_key]["handoffs"] += 1
            if row.source == "guardrail":
                daily_map[day_key]["safety"] += 1
            if row.processing_status == "failed" or bool(decrypt_if_needed(row.error)):
                daily_map[day_key]["failed"] += 1
            if _is_homepage_request(decrypt_if_needed(row.question)):
                daily_map[day_key]["homepage_requests"] += 1

    for row in analysis_cancels:
        if not row.created_at:
            continue
        month_key = _analysis_datetime(row.created_at).strftime("%Y-%m")
        metric_key = "refunds" if _is_refund_request(decrypt_if_needed(row.message)) else "cancels"
        if month_key in monthly_map:
            monthly_map[month_key][metric_key] += 1
        day_key = _analysis_datetime(row.created_at).date().isoformat()
        if day_key in daily_map:
            daily_map[day_key][metric_key] += 1

    question_category_map: dict[str, dict] = {}
    answer_source_summary = {"faq": 0, "llm": 0, "other": 0}
    handoff_category_map: dict[str, dict] = {}
    cancel_keys = {
        (row.session_id, (decrypt_if_needed(row.message) or "").strip())
        for row in analysis_cancels
    }

    session_context_map: dict[str, list[str]] = {}
    for context_row in analysis_logs:
        session_context_map.setdefault(context_row.session_id, []).extend([
            decrypt_if_needed(context_row.question) or "",
            decrypt_if_needed(context_row.answer) or "",
        ])

    unclassified_count = 0
    for row in analysis_logs:
        question = decrypt_if_needed(row.question) or ""
        if row.question_category and row.question_category != "general":
            category_key = row.question_category
            category_label = row.question_category_label or row.question_category
        else:
            category_key, category_label = _question_category(question)
        if category_key == "general":
            unclassified_count += 1
        category = question_category_map.setdefault(
            category_key,
            {"key": category_key, "label": category_label, "count": 0},
        )
        category["count"] += 1

        if row.source == "faq":
            answer_source_summary["faq"] += 1
        elif row.source in {"ai", "document"}:
            answer_source_summary["llm"] += 1
        else:
            answer_source_summary["other"] += 1

        is_handoff = row.processing_status in ("handoff", "handoff_offer") or row.source == "handoff"
        if is_handoff:
            is_cancel = (row.session_id, question.strip()) in cancel_keys
            context = " ".join(session_context_map.get(row.session_id, []))
            handoff_key, handoff_label = _handoff_reason(
                question,
                row.processing_status,
                is_cancel,
                decrypt_if_needed(row.answer) or "",
                context,
            )
            handoff_category = handoff_category_map.setdefault(
                handoff_key,
                {"key": handoff_key, "label": handoff_label, "count": 0},
            )
            handoff_category["count"] += 1

    monthly = list(monthly_map.values())
    hourly = list(hourly_map.values())
    question_categories_top5 = sorted(
        (item for item in question_category_map.values() if item["key"] != "general"),
        key=lambda item: (-item["count"], item["label"]),
    )[:5]
    handoff_categories = list(handoff_category_map.values())
    handoff_categories.sort(key=lambda item: item["count"], reverse=True)
    chat_count = len(analysis_logs)
    visitor_count = len(analysis_sessions)
    handoff_count = sum(item["count"] for item in handoff_categories)
    failed_count = sum(
        1 for row in analysis_logs
        if row.processing_status == "failed" or bool(decrypt_if_needed(row.error))
    )
    refund_count = sum(1 for row in analysis_cancels if _is_refund_request(decrypt_if_needed(row.message)))
    homepage_request_count = sum(
        1 for row in analysis_logs if _is_homepage_request(decrypt_if_needed(row.question))
    )
    period_summary = {
        "visitors": visitor_count,
        "chats": chat_count,
        "handoffs": handoff_count,
        "cancels": len(analysis_cancels) - refund_count,
        "refunds": refund_count,
        "homepage_requests": homepage_request_count,
        "safety": sum(1 for row in analysis_logs if row.source == "guardrail"),
        "failed": failed_count,
    }

    return {
        "period_months": len(month_starts),
        "hourly_days": max(1, (analysis_end.date() - analysis_start.date()).days),
        "selected_year": selected_year,
        "selected_month": selected_month,
        "available_years": list(range(first_available_year, last_available_year + 1)),
        "available_months": available_months,
        "period_label": (
            f"{selected_start.year}년 {selected_start.month}월"
            if selected_month
            else f"{selected_year}년" if selected_year is not None else "전체 기간"
        ),
        "generated_at": datetime.now(),
        "monthly": monthly,
        "daily": list(daily_map.values()),
        "hourly": hourly,
        "period_summary": period_summary,
        "highlights": {
            "busiest_visitor_month": _peak_item(monthly, "visitors", "month"),
            "busiest_chat_month": _peak_item(monthly, "chats", "month"),
            "busiest_visitor_hour": _peak_item(hourly, "visitors", "label"),
            "busiest_chat_hour": _peak_item(hourly, "chats", "label"),
        },
        "question_categories_top5": question_categories_top5,
        "answer_source_summary": {
            **answer_source_summary,
            "total": sum(answer_source_summary.values()),
        },
        "handoff_categories": handoff_categories,
        "unclassified_count": unclassified_count,
    }


@router.post("/operations/analytics/reclassify")
async def reclassify_question_categories(
    limit: int = Query(500, ge=1, le=2000),
    db: Session = Depends(get_db),
    current_user: str = Depends(verify_admin),
):
    rows = db.query(ChatLog).filter(
        or_(
            ChatLog.question_category.is_(None),
            ChatLog.question_category == "general",
            ChatLog.question_category_source == "pending",
        )
    ).order_by(ChatLog.created_at.desc()).limit(limit).all()

    pending_items = []
    rule_classified = 0
    for row in rows:
        question = decrypt_if_needed(row.question) or ""
        category = categorize_question_rule(question)
        if category.key != "general":
            row.question_category = category.key
            row.question_category_label = category.label
            row.question_category_source = category.source
            rule_classified += 1
        else:
            pending_items.append({"id": row.id, "question": question})

    llm_classified = 0
    for offset in range(0, len(pending_items), 50):
        classified = await classify_questions_batch(pending_items[offset:offset + 50])
        if not classified:
            continue
        batch_ids = list(classified.keys())
        batch_rows = db.query(ChatLog).filter(ChatLog.id.in_(batch_ids)).all()
        for row in batch_rows:
            category = classified.get(row.id)
            if category is None:
                continue
            row.question_category = category.key
            row.question_category_label = category.label
            row.question_category_source = category.source
            llm_classified += 1

    db.commit()
    remaining = max(0, len(pending_items) - llm_classified)
    create_audit_log(
        db,
        "question_categories_reclassified",
        "chat_log",
        None,
        f"rule={rule_classified}, llm={llm_classified}, remaining={remaining}",
        actor=current_user,
    )
    return {
        "classified": rule_classified + llm_classified,
        "rule_classified": rule_classified,
        "llm_classified": llm_classified,
        "remaining": remaining,
    }


@router.get("/operations/health")
def get_operations_health(_: None = Depends(verify_admin)):
    """Check API availability, a real application-table read, a committed write, and EC2 status."""
    checked_at = datetime.now()
    checks = [{
        "key": "application",
        "label": "백엔드 API",
        "status": "healthy",
        "message": "관리자 API가 정상 응답 중입니다.",
        "latency_ms": 0,
        "checked_at": checked_at,
        "details": {},
    }]

    read_started = perf_counter()
    read_db = SessionLocal()
    try:
        read_db.query(ChatLog.id).order_by(ChatLog.id.desc()).limit(1).first()
        checks.append({
            "key": "database_read",
            "label": "DB 조회",
            "status": "healthy",
            "message": "채팅 데이터 조회가 정상입니다.",
            "latency_ms": round((perf_counter() - read_started) * 1000),
            "checked_at": checked_at,
            "details": {},
        })
    except Exception as exc:
        checks.append({
            "key": "database_read",
            "label": "DB 조회",
            "status": "critical",
            "message": "데이터를 조회하지 못했습니다.",
            "latency_ms": round((perf_counter() - read_started) * 1000),
            "checked_at": checked_at,
            "details": {"error_type": type(exc).__name__},
        })
    finally:
        read_db.close()

    write_started = perf_counter()
    write_db = SessionLocal()
    probe_nonce = uuid4().hex
    try:
        probe = write_db.get(SystemHealthProbe, "database_write")
        if probe is None:
            probe = SystemHealthProbe(key="database_write", nonce=probe_nonce, checked_at=checked_at)
            write_db.add(probe)
        else:
            probe.nonce = probe_nonce
            probe.checked_at = checked_at
        write_db.commit()
        write_db.close()

        verify_db = SessionLocal()
        persisted_probe = verify_db.get(SystemHealthProbe, "database_write")
        if persisted_probe is None or persisted_probe.nonce != probe_nonce:
            raise RuntimeError("database write verification failed")
        verify_db.close()
        checks.append({
            "key": "database_write",
            "label": "DB 저장",
            "status": "healthy",
            "message": "테스트 데이터 저장과 재조회가 정상입니다.",
            "latency_ms": round((perf_counter() - write_started) * 1000),
            "checked_at": checked_at,
            "details": {"last_success_at": checked_at},
        })
    except Exception as exc:
        write_db.rollback()
        write_db.close()
        checks.append({
            "key": "database_write",
            "label": "DB 저장",
            "status": "critical",
            "message": "데이터 저장 또는 저장 결과 확인에 실패했습니다.",
            "latency_ms": round((perf_counter() - write_started) * 1000),
            "checked_at": checked_at,
            "details": {"error_type": type(exc).__name__},
        })

    checks.append(_ec2_health_check())
    statuses = {check["status"] for check in checks}
    overall_status = "critical" if "critical" in statuses else "degraded" if "unknown" in statuses else "healthy"
    return {
        "overall_status": overall_status,
        "generated_at": checked_at,
        "checks": checks,
    }


@router.get("/operations/dashboard")
def get_operations_dashboard(
    days: int = Query(7, ge=1, le=30),
    attention_limit: int = Query(50, ge=10, le=200),
    db: Session = Depends(get_db),
    _: None = Depends(verify_admin),
):
    """운영 대시보드용 지표, 상담 사유, 주의 대화를 한 번에 반환한다."""
    today = datetime.now().date()
    start_date = today - timedelta(days=days - 1)
    previous_start_date = start_date - timedelta(days=days)
    previous_start = datetime.combine(previous_start_date, time.min)

    sessions = db.query(ChatSession).filter(ChatSession.created_at >= previous_start).all()
    logs = db.query(ChatLog).filter(ChatLog.created_at >= previous_start).order_by(ChatLog.created_at.desc()).all()
    cancels = db.query(CancelRequest).filter(CancelRequest.created_at >= previous_start).order_by(CancelRequest.created_at.desc()).all()

    current_sessions = [row for row in sessions if row.created_at and row.created_at.date() >= start_date]
    previous_sessions = [row for row in sessions if row.created_at and previous_start_date <= row.created_at.date() < start_date]
    current_logs = [row for row in logs if row.created_at and row.created_at.date() >= start_date]
    previous_logs = [row for row in logs if row.created_at and previous_start_date <= row.created_at.date() < start_date]
    current_cancels = [row for row in cancels if row.created_at and row.created_at.date() >= start_date]
    previous_cancels = [row for row in cancels if row.created_at and previous_start_date <= row.created_at.date() < start_date]

    summary = _operations_summary(current_sessions, current_logs, current_cancels)
    previous_summary = _operations_summary(previous_sessions, previous_logs, previous_cancels)
    changes = {key: _percent_change(summary[key], previous_summary[key]) for key in summary}

    daily_map = {
        (start_date + timedelta(days=offset)).isoformat(): {
            "date": (start_date + timedelta(days=offset)).isoformat(),
            "visitors": 0,
            "chats": 0,
            "handoffs": 0,
            "cancels": 0,
            "refunds": 0,
            "homepage_requests": 0,
            "safety": 0,
            "failed": 0,
        }
        for offset in range(days)
    }
    for row in current_sessions:
        daily_map[row.created_at.date().isoformat()]["visitors"] += 1
    for row in current_logs:
        bucket = daily_map[row.created_at.date().isoformat()]
        bucket["chats"] += 1
        if row.processing_status == "handoff" or row.source == "handoff":
            bucket["handoffs"] += 1
        if row.source == "guardrail":
            bucket["safety"] += 1
        if row.processing_status == "failed" or bool(decrypt_if_needed(row.error)):
            bucket["failed"] += 1
        if _is_homepage_request(decrypt_if_needed(row.question)):
            bucket["homepage_requests"] += 1
    for row in current_cancels:
        key = "refunds" if _is_refund_request(decrypt_if_needed(row.message)) else "cancels"
        daily_map[row.created_at.date().isoformat()][key] += 1

    request_signal_by_key = {
        (row.session_id, (decrypt_if_needed(row.message) or "").strip()): (
            "refund" if _is_refund_request(decrypt_if_needed(row.message)) else "cancel"
        )
        for row in current_cancels
    }
    handoff_category_map: dict[str, dict] = {}
    session_context_map: dict[str, list[str]] = {}
    for context_row in current_logs:
        session_context_map.setdefault(context_row.session_id, []).extend([
            decrypt_if_needed(context_row.question) or "",
            decrypt_if_needed(context_row.answer) or "",
        ])
    attention = []
    log_ids = [row.id for row in current_logs]
    alert_by_log_id = {
        alert.chat_log_id: alert
        for alert in db.query(OperationsAlert).filter(OperationsAlert.chat_log_id.in_(log_ids)).all()
    } if log_ids else {}
    alerts_changed = False

    for row in current_logs:
        question = decrypt_if_needed(row.question) or ""
        answer = decrypt_if_needed(row.answer) or ""
        error = decrypt_if_needed(row.error)
        request_signal = request_signal_by_key.get((row.session_id, question.strip()))
        is_cancel = request_signal is not None
        is_handoff = row.processing_status in ("handoff", "handoff_offer") or row.source == "handoff"
        reason = ""
        if is_handoff:
            category_key, reason = _handoff_reason(
                question,
                row.processing_status,
                is_cancel,
                answer,
                " ".join(session_context_map.get(row.session_id, [])),
            )
            category = handoff_category_map.setdefault(
                category_key,
                {"key": category_key, "label": reason, "count": 0},
            )
            category["count"] += 1

        signal_type = None
        severity = "medium"
        if request_signal == "refund":
            signal_type, severity, reason = "refund", "medium", "환불 요청 접수"
        elif request_signal == "cancel":
            signal_type, severity, reason = "cancel", "medium", "취소 요청 접수"
        elif row.source == "guardrail":
            signal_type, severity, reason = "safety", "high", "안전 가드레일 감지"
        elif row.processing_status == "failed" or error:
            signal_type, severity, reason = "error", "high", "응답 처리 오류"
        elif is_handoff:
            signal_type = "handoff"
            severity = "low" if row.processing_status == "handoff_offer" else "medium"

        if signal_type:
            alert = alert_by_log_id.get(row.id)
            if alert is None:
                alert = OperationsAlert(
                    chat_log_id=row.id,
                    session_id=row.session_id,
                    signal_type=signal_type,
                    severity=severity,
                    reason=reason,
                    status="open",
                )
                db.add(alert)
                db.flush()
                alert_by_log_id[row.id] = alert
                alerts_changed = True
            elif alert.signal_type != signal_type or alert.severity != severity or alert.reason != reason:
                alert.signal_type = signal_type
                alert.severity = severity
                alert.reason = reason
                alerts_changed = True

        if signal_type and len(attention) < attention_limit:
            attention.append({
                "id": row.id,
                "alert_id": alert.id,
                "session_id": row.session_id,
                "type": signal_type,
                "severity": severity,
                "reason": reason,
                "status": alert.status,
                "assigned_to": alert.assigned_to,
                "note": decrypt_if_needed(alert.note),
                "question": question,
                "answer": answer,
                "processing_status": row.processing_status,
                "created_at": row.created_at,
            })

    if alerts_changed:
        db.commit()

    handoff_categories = list(handoff_category_map.values())
    handoff_categories.sort(key=lambda item: item["count"], reverse=True)

    question_category_map: dict[str, dict] = {}
    answer_source_summary = {"faq": 0, "llm": 0, "other": 0}
    for row in current_logs:
        if row.question_category and row.question_category != "general":
            category_key = row.question_category
            category_label = row.question_category_label or row.question_category
        else:
            category_key, category_label = _question_category(decrypt_if_needed(row.question) or "")
        category = question_category_map.setdefault(
            category_key,
            {"key": category_key, "label": category_label, "count": 0},
        )
        category["count"] += 1

        if row.source == "faq":
            answer_source_summary["faq"] += 1
        elif row.source in {"ai", "document"}:
            answer_source_summary["llm"] += 1
        else:
            answer_source_summary["other"] += 1

    question_categories_top5 = sorted(
        (item for item in question_category_map.values() if item["key"] != "general"),
        key=lambda item: (-item["count"], item["label"]),
    )[:5]
    recent_sessions = sorted(
        current_sessions,
        key=lambda row: row.updated_at or row.created_at or datetime.min,
        reverse=True,
    )[:10]
    last_conversation_at = db.query(func.max(ChatMessage.created_at)).scalar()
    return {
        "period_days": days,
        "generated_at": datetime.now(),
        "last_conversation_at": last_conversation_at,
        "summary": summary,
        "previous_summary": previous_summary,
        "changes": changes,
        "daily": list(daily_map.values()),
        "handoff_categories": handoff_categories,
        "question_categories_top5": question_categories_top5,
        "answer_source_summary": {
            **answer_source_summary,
            "total": sum(answer_source_summary.values()),
        },
        "attention": attention,
        "recent_sessions": [
            {
                "id": row.id,
                "created_at": row.created_at,
                "updated_at": row.updated_at,
                "message_count": row.message_count,
                "user_name": decrypt_if_needed(row.encrypted_user_name) if row.encrypted_user_name else None,
            }
            for row in recent_sessions
        ],
    }


def _find_openai_project_id(admin_key: str, project_name: str) -> str:
    after: str | None = None
    matches: list[str] = []
    while True:
        params: dict[str, object] = {"limit": 100, "include_archived": "false"}
        if after:
            params["after"] = after
        request = Request(
            f"https://api.openai.com/v1/organization/projects?{urlencode(params)}",
            headers={"Authorization": f"Bearer {admin_key}", "Accept": "application/json"},
        )
        with urlopen(request, timeout=20) as response:
            payload = json.loads(response.read().decode("utf-8"))
        matches.extend(
            str(project["id"])
            for project in payload.get("data", [])
            if str(project.get("name") or "").strip().casefold() == project_name.casefold()
            and project.get("status") == "active"
        )
        if not payload.get("has_more"):
            break
        after = str(payload.get("last_id") or "")
        if not after:
            break
    if not matches:
        raise LookupError(f"OpenAI 프로젝트 '{project_name}'을 찾지 못했습니다.")
    if len(matches) > 1:
        raise LookupError(f"이름이 같은 OpenAI 프로젝트가 여러 개입니다. OPENAI_PROJECT_ID를 설정해 주세요.")
    return matches[0]


@router.get("/operations/openai-costs")
def get_openai_costs(
    billing_month: str = Query(..., pattern=r"^\d{4}-\d{2}$"),
    _: None = Depends(verify_admin),
):
    settings = get_settings()
    fetched_at = datetime.now(ZoneInfo("Asia/Seoul")).isoformat()
    project_name = settings.openai_project_name.strip() or "AIcampus_Chatbot"
    if not settings.openai_admin_key:
        return {
            "billing_month": billing_month,
            "project_id": None,
            "project_name": project_name,
            "configured": False,
            "status": "not_configured",
            "message": f"OPENAI_ADMIN_KEY를 설정하면 {project_name} 프로젝트 비용을 자동으로 가져옵니다.",
            "currency": "usd",
            "total_usd": 0,
            "daily": [],
            "line_items": [],
            "fetched_at": fetched_at,
        }

    try:
        month_start = datetime.strptime(billing_month, "%Y-%m").date().replace(day=1)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="청구 월은 YYYY-MM 형식이어야 합니다.") from exc
    month_end = _shift_month(month_start, 1)
    start_time = int(datetime.combine(month_start, time.min, tzinfo=ZoneInfo("UTC")).timestamp())
    end_time = int(datetime.combine(month_end, time.min, tzinfo=ZoneInfo("UTC")).timestamp())

    daily_map: dict[str, float] = {}
    line_item_map: dict[str, float] = {}
    currency = "usd"
    page: str | None = None

    try:
        project_id = settings.openai_project_id.strip() or _find_openai_project_id(
            settings.openai_admin_key,
            project_name,
        )
        while True:
            params: dict[str, object] = {
                "start_time": start_time,
                "end_time": end_time,
                "bucket_width": "1d",
                "limit": 31,
                "group_by": ["line_item"],
            }
            params["project_ids"] = [project_id]
            if page:
                params["page"] = page
            request = Request(
                f"https://api.openai.com/v1/organization/costs?{urlencode(params, doseq=True)}",
                headers={
                    "Authorization": f"Bearer {settings.openai_admin_key}",
                    "Accept": "application/json",
                },
            )
            with urlopen(request, timeout=20) as response:
                payload = json.loads(response.read().decode("utf-8"))

            for bucket in payload.get("data", []):
                bucket_date = datetime.fromtimestamp(bucket["start_time"], tz=ZoneInfo("UTC")).date().isoformat()
                for result in bucket.get("results", []):
                    amount = result.get("amount") or {}
                    value = float(amount.get("value") or 0)
                    currency = str(amount.get("currency") or currency).lower()
                    daily_map[bucket_date] = daily_map.get(bucket_date, 0) + value
                    line_item = str(result.get("line_item") or "기타")
                    line_item_map[line_item] = line_item_map.get(line_item, 0) + value

            if not payload.get("has_more") or not payload.get("next_page"):
                break
            page = str(payload["next_page"])
    except LookupError as exc:
        return {
            "billing_month": billing_month,
            "project_id": None,
            "project_name": project_name,
            "configured": True,
            "status": "error",
            "message": str(exc),
            "currency": "usd",
            "total_usd": 0,
            "daily": [],
            "line_items": [],
            "fetched_at": fetched_at,
        }
    except HTTPError as exc:
        message = "OpenAI 비용 조회 권한을 확인해 주세요." if exc.code in (401, 403) else "OpenAI 비용 데이터를 불러오지 못했습니다."
        return {
            "billing_month": billing_month,
            "project_id": None,
            "project_name": project_name,
            "configured": True,
            "status": "error",
            "message": message,
            "currency": "usd",
            "total_usd": 0,
            "daily": [],
            "line_items": [],
            "fetched_at": fetched_at,
        }
    except (URLError, TimeoutError, ValueError, KeyError, json.JSONDecodeError):
        return {
            "billing_month": billing_month,
            "project_id": None,
            "project_name": project_name,
            "configured": True,
            "status": "error",
            "message": "OpenAI 비용 데이터를 불러오지 못했습니다.",
            "currency": "usd",
            "total_usd": 0,
            "daily": [],
            "line_items": [],
            "fetched_at": fetched_at,
        }

    daily = [{"date": key, "amount_usd": round(value, 6)} for key, value in sorted(daily_map.items())]
    line_items = [
        {"line_item": key, "amount_usd": round(value, 6)}
        for key, value in sorted(line_item_map.items(), key=lambda item: item[1], reverse=True)
    ]
    return {
        "billing_month": billing_month,
        "project_id": project_id,
        "project_name": project_name,
        "configured": True,
        "status": "available",
        "message": f"{project_name} 프로젝트 청구 비용 기준입니다.",
        "currency": currency,
        "total_usd": round(sum(daily_map.values()), 6),
        "daily": daily,
        "line_items": line_items,
        "fetched_at": fetched_at,
    }


@router.get("/operations/cost-management")
def get_cost_management(
    billing_month: str = Query(...),
    account_id: str = Query("all"),
    db: Session = Depends(get_db),
    _: None = Depends(verify_admin),
):
    is_all_period = billing_month == "all"
    month_start: date | None = None
    month_end: date | None = None
    if not is_all_period:
        try:
            month_start = datetime.strptime(billing_month, "%Y-%m").date().replace(day=1)
        except ValueError:
            raise HTTPException(status_code=400, detail="청구 월은 YYYY-MM 형식 또는 all이어야 합니다.")
        month_end = _shift_month(month_start, 1)

    rows_query = db.query(BillingDailyCostRecord)
    if month_start is not None and month_end is not None:
        rows_query = rows_query.filter(
            BillingDailyCostRecord.usage_date >= month_start,
            BillingDailyCostRecord.usage_date < month_end,
        )
    all_rows = rows_query.all()

    account_map: dict[str, dict] = {}
    for row in all_rows:
        account = account_map.setdefault(
            row.account_id,
            {"account_id": row.account_id, "account_name": row.account_name, "total_krw": 0},
        )
        account["total_krw"] += row.amount_krw
    accounts = sorted(account_map.values(), key=lambda item: item["account_name"])
    selected_rows = all_rows if account_id == "all" else [row for row in all_rows if row.account_id == account_id]

    if month_start is not None:
        _, last_day = calendar.monthrange(month_start.year, month_start.month)
        day_keys = [f"{day:02d}" for day in range(1, last_day + 1)]
    else:
        day_keys = []
    service_map: dict[str, dict] = {}
    daily_map = {
        day: {"date": f"{billing_month}-{day}", "day": int(day), "total_krw": 0, "services": {}}
        for day in day_keys
    }
    for row in selected_rows:
        service = service_map.setdefault(
            row.service_name,
            {"service_name": row.service_name, "total_krw": 0, "daily": {day: 0 for day in day_keys}},
        )
        service["total_krw"] += row.amount_krw
        if not is_all_period:
            day_key = f"{row.usage_date.day:02d}"
            service["daily"][day_key] += row.amount_krw
            daily_map[day_key]["total_krw"] += row.amount_krw
            daily_map[day_key]["services"][row.service_name] = (
                daily_map[day_key]["services"].get(row.service_name, 0) + row.amount_krw
            )

    services = sorted(service_map.values(), key=lambda item: item["total_krw"], reverse=True)
    history_query = db.query(BillingDailyCostRecord)
    if account_id != "all":
        history_query = history_query.filter(BillingDailyCostRecord.account_id == account_id)
    monthly_map: dict[str, int] = {}
    for row in history_query.all():
        month_key = row.usage_date.strftime("%Y-%m")
        monthly_map[month_key] = monthly_map.get(month_key, 0) + row.amount_krw
    monthly_history = [
        {"billing_month": month, "amount_krw": amount}
        for month, amount in sorted(monthly_map.items())
    ]

    uploaded_file = None
    if not is_all_period:
        upload_account_id = NXAVIS_ACCOUNT_ID if account_id == "all" else account_id
        uploaded_file = db.query(BillingCostUploadRecord).filter(
            BillingCostUploadRecord.billing_month == billing_month,
            BillingCostUploadRecord.account_id == upload_account_id,
        ).first()
    usage_total = sum(row.amount_krw for row in selected_rows)
    return {
        "billing_month": billing_month,
        "is_all_period": is_all_period,
        "selected_account_id": account_id,
        "accounts": accounts,
        "usage_total_krw": usage_total,
        "service_totals": [
            {"service_name": item["service_name"], "amount_krw": item["total_krw"]}
            for item in services
        ],
        "daily_totals": list(daily_map.values()),
        "service_daily_rows": services,
        "monthly_history": monthly_history,
        "uploaded_file": _serialize_billing_cost_upload(uploaded_file),
    }


@router.get("/operations/costs/template")
def download_cost_template(
    billing_month: str | None = Query(None, pattern=r"^\d{4}-\d{2}$"),
    _: None = Depends(verify_admin),
):
    template_month = billing_month or datetime.now().strftime("%Y-%m")
    try:
        datetime.strptime(template_month, "%Y-%m")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="유효한 반영 월을 선택해 주세요.") from exc
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "일별 서비스 비용"
    sheet.append(["사용일자", "계정ID", "계정명", "서비스", "원화금액"])
    sheet.append([f"{template_month}-01", NXAVIS_ACCOUNT_ID, NXAVIS_ACCOUNT_NAME, "AmazonRDS", 0])
    buffer = io.BytesIO()
    workbook.save(buffer)
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="cost_import_template_{template_month}.xlsx"'},
    )


@router.get("/operations/costs/uploaded-file/{billing_month}")
def download_uploaded_cost_file(
    billing_month: str,
    account_id: str = Query(NXAVIS_ACCOUNT_ID),
    db: Session = Depends(get_db),
    _: None = Depends(verify_admin),
):
    try:
        parsed_month = datetime.strptime(billing_month, "%Y-%m")
        if parsed_month.strftime("%Y-%m") != billing_month:
            raise ValueError
    except ValueError:
        raise HTTPException(status_code=400, detail="청구 월은 YYYY-MM 형식이어야 합니다.")

    upload = db.query(BillingCostUploadRecord).filter(
        BillingCostUploadRecord.billing_month == billing_month,
        BillingCostUploadRecord.account_id == account_id,
    ).first()
    if upload is None:
        raise HTTPException(status_code=404, detail="해당 월에 업로드된 비용 파일이 없습니다.")

    fallback_filename = re.sub(r"[^A-Za-z0-9._-]", "_", upload.original_filename) or f"costs_{billing_month}"
    encoded_filename = quote(upload.original_filename)
    return Response(
        content=upload.file_data,
        media_type=upload.content_type,
        headers={
            "Content-Disposition": (
                f'attachment; filename="{fallback_filename}"; filename*=UTF-8\'\'{encoded_filename}'
            )
        },
    )


@router.post("/operations/costs/import")
async def import_cost_details(
    file: UploadFile = File(...),
    billing_month: str = Form(""),
    account_id: str = Form(""),
    account_name: str = Form(""),
    db: Session = Depends(get_db),
    current_user: str = Depends(verify_admin),
):
    filename = file.filename or ""
    safe_filename = Path(filename).name.replace("\r", "").replace("\n", "")[:255]
    raw = await file.read()
    if filename.lower().endswith(".csv"):
        decoded = raw.decode("utf-8-sig")
        values = list(csv.reader(io.StringIO(decoded)))
    elif filename.lower().endswith(".xlsx"):
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        values = [list(row) for row in workbook.active.iter_rows(values_only=True)]
    else:
        raise HTTPException(status_code=400, detail=".xlsx 또는 .csv 파일만 업로드할 수 있습니다.")
    if len(values) < 2:
        raise HTTPException(status_code=400, detail="비용 데이터 행이 없습니다.")
    if not re.fullmatch(r"\d{4}-\d{2}", billing_month):
        raise HTTPException(status_code=400, detail="반영 월은 YYYY-MM 형식이어야 합니다.")
    try:
        datetime.strptime(billing_month, "%Y-%m")
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="유효한 반영 월을 선택해 주세요.") from exc

    raw_headers = values[0]
    headers = [_normalize_cost_header(value) for value in raw_headers]
    aliases = {
        "date": {"사용일자", "사용일", "일자", "date", "usagedate"},
        "account_id": {"계정id", "계정번호", "accountid", "account"},
        "account_name": {"계정명", "accountname"},
        "service": {"서비스", "서비스명", "service", "servicename"},
        "amount": {"원화금액", "사용금액", "금액", "amount", "amountkrw", "krw"},
    }

    def column_index(key: str) -> int | None:
        return next((index for index, header in enumerate(headers) if header in aliases[key]), None)

    date_index = column_index("date")
    service_index = column_index("service")
    amount_index = column_index("amount")
    parsed_rows: list[tuple[date, str, str, str, int]] = []

    try:
        if date_index is not None and service_index is not None and amount_index is not None:
            account_id_index = column_index("account_id")
            account_name_index = column_index("account_name")
            for values_row in values[1:]:
                if not any(value not in (None, "") for value in values_row):
                    continue
                usage_date = _parse_cost_date(values_row[date_index])
                row_account_id = str(values_row[account_id_index]).strip() if account_id_index is not None else account_id.strip()
                row_account_name = str(values_row[account_name_index]).strip() if account_name_index is not None else account_name.strip()
                service_name = str(values_row[service_index] or "").strip()
                amount_krw = _parse_krw_amount(values_row[amount_index])
                if not row_account_id or not row_account_name or not service_name:
                    raise ValueError("계정 또는 서비스 값 누락")
                parsed_rows.append((usage_date, row_account_id, row_account_name, service_name, amount_krw))
        else:
            if not re.fullmatch(r"\d{4}-\d{2}", billing_month) or not account_id.strip() or not account_name.strip():
                raise ValueError("행렬 형식은 청구 월과 계정 정보가 필요합니다")
            service_index = service_index if service_index is not None else 0
            day_columns = []
            for index, header in enumerate(headers):
                match = re.fullmatch(r"(\d{1,2})일?", header)
                if match and 1 <= int(match.group(1)) <= 31:
                    day_columns.append((index, int(match.group(1))))
            if not day_columns:
                raise ValueError("일자 열을 찾지 못했습니다")
            for values_row in values[1:]:
                service_name = str(values_row[service_index] or "").strip()
                if not service_name or service_name.lower() in {"total", "합계"}:
                    continue
                for index, day in day_columns:
                    amount_krw = _parse_krw_amount(values_row[index] if index < len(values_row) else 0)
                    if amount_krw == 0:
                        continue
                    usage_date = datetime.strptime(f"{billing_month}-{day:02d}", "%Y-%m-%d").date()
                    parsed_rows.append((usage_date, account_id.strip(), account_name.strip(), service_name, amount_krw))
    except (ValueError, TypeError, IndexError) as exc:
        raise HTTPException(status_code=400, detail=f"비용 파일 형식을 확인하세요: {exc}")

    mismatched_dates = sorted({
        usage_date.strftime("%Y-%m-%d")
        for usage_date, *_ in parsed_rows
        if usage_date.strftime("%Y-%m") != billing_month
    })
    if mismatched_dates:
        preview = ", ".join(mismatched_dates[:3])
        raise HTTPException(
            status_code=400,
            detail=f"선택한 반영 월({billing_month})과 다른 날짜가 포함되어 있습니다: {preview}",
        )

    if not parsed_rows:
        raise HTTPException(status_code=400, detail="반영할 비용 항목이 없습니다.")

    aggregated_rows: dict[tuple[date, str], int] = {}
    for usage_date, _, _, service_name, amount_krw in parsed_rows:
        key = (usage_date, service_name)
        aggregated_rows[key] = aggregated_rows.get(key, 0) + amount_krw

    month_start = datetime.strptime(billing_month, "%Y-%m").date().replace(day=1)
    month_end = _shift_month(month_start, 1)
    replaced_rows = db.query(BillingDailyCostRecord).filter(
        BillingDailyCostRecord.usage_date >= month_start,
        BillingDailyCostRecord.usage_date < month_end,
        BillingDailyCostRecord.account_id == NXAVIS_ACCOUNT_ID,
    ).delete(synchronize_session=False)

    for (usage_date, service_name), amount_krw in aggregated_rows.items():
        db.add(BillingDailyCostRecord(
            usage_date=usage_date,
            account_id=NXAVIS_ACCOUNT_ID,
            account_name=NXAVIS_ACCOUNT_NAME,
            service_name=service_name,
            amount_krw=amount_krw,
            source="nxavis_excel",
        ))

    upload = db.query(BillingCostUploadRecord).filter(
        BillingCostUploadRecord.billing_month == billing_month,
        BillingCostUploadRecord.account_id == NXAVIS_ACCOUNT_ID,
    ).first()
    if upload is None:
        upload = BillingCostUploadRecord(
            billing_month=billing_month,
            account_id=NXAVIS_ACCOUNT_ID,
        )
        db.add(upload)
    upload.account_name = NXAVIS_ACCOUNT_NAME
    upload.original_filename = safe_filename or f"costs_{billing_month}{Path(filename).suffix.lower()}"
    upload.content_type = file.content_type or (
        "text/csv" if filename.lower().endswith(".csv")
        else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    upload.file_data = raw
    upload.size_bytes = len(raw)
    upload.imported_rows = len(aggregated_rows)
    upload.uploaded_by = current_user
    db.commit()
    create_audit_log(
        db,
        "billing_cost_imported",
        "billing_daily_cost",
        filename,
        f"rows={len(aggregated_rows)}, replaced_rows={replaced_rows}",
        actor=current_user,
    )
    return {
        "message": f"{billing_month} 실제 원화 비용 자료를 반영했습니다.",
        "billing_month": billing_month,
        "imported_rows": len(aggregated_rows),
        "replaced_rows": replaced_rows,
        "filename": upload.original_filename,
    }


def _add_operations_alert_history(
    db: Session,
    alert: OperationsAlert,
    action: str,
    actor: str,
    from_status: str | None = None,
) -> None:
    db.add(OperationsAlertHistory(
        alert_id=alert.id,
        action=action,
        actor=actor,
        from_status=from_status,
        to_status=alert.status,
        test_question=alert.test_question,
        test_answer=alert.test_answer,
        test_source=alert.test_source,
        test_passed=alert.test_passed,
    ))


def _serialize_operations_alert(alert: OperationsAlert) -> dict:
    return {
        "id": alert.id,
        "session_id": alert.session_id,
        "signal_type": alert.signal_type,
        "severity": alert.severity,
        "reason": alert.reason,
        "status": alert.status,
        "assigned_to": alert.assigned_to,
        "note": decrypt_if_needed(alert.note),
        "test_question": decrypt_if_needed(alert.test_question),
        "test_answer": decrypt_if_needed(alert.test_answer),
        "test_source": alert.test_source,
        "test_passed": bool(alert.test_passed),
        "tested_by": alert.tested_by,
        "tested_at": alert.tested_at,
        "created_at": alert.created_at,
        "updated_at": alert.updated_at,
        "resolved_at": alert.resolved_at,
    }


@router.get("/operations/alerts/{alert_id}/detail")
def get_operations_alert_detail(
    alert_id: int,
    db: Session = Depends(get_db),
    _: str = Depends(verify_admin),
):
    alert = db.query(OperationsAlert).filter(OperationsAlert.id == alert_id).first()
    if not alert:
        raise HTTPException(status_code=404, detail="운영 알림을 찾을 수 없습니다.")
    messages = db.query(ChatMessage).filter(
        ChatMessage.session_id == alert.session_id
    ).order_by(ChatMessage.created_at, ChatMessage.id).all()
    source_log = db.query(ChatLog).filter(ChatLog.id == alert.chat_log_id).first()
    problem_question = (decrypt_if_needed(source_log.question) or "").strip() if source_log else ""
    problem_start_message_id = None
    for message in messages:
        if message.role == "user" and (decrypt_if_needed(message.content) or "").strip() == problem_question:
            problem_start_message_id = message.id
    if problem_start_message_id is None and source_log and source_log.created_at:
        for message in messages:
            if message.role == "user" and message.created_at and message.created_at <= source_log.created_at:
                problem_start_message_id = message.id

    histories = db.query(OperationsAlertHistory).filter(
        OperationsAlertHistory.alert_id == alert.id
    ).order_by(OperationsAlertHistory.created_at.desc(), OperationsAlertHistory.id.desc()).all()
    related_alerts = db.query(OperationsAlert).filter(
        OperationsAlert.id != alert.id,
        OperationsAlert.status == "resolved",
        OperationsAlert.signal_type == alert.signal_type,
        OperationsAlert.reason == alert.reason,
    ).order_by(OperationsAlert.resolved_at.desc(), OperationsAlert.id.desc()).limit(5).all()
    return {
        "alert": _serialize_operations_alert(alert),
        "problem_start_message_id": problem_start_message_id,
        "messages": [
            {
                "id": message.id,
                "role": message.role,
                "content": decrypt_if_needed(message.content) or "",
                "source": message.source,
                "created_at": message.created_at,
            }
            for message in messages
        ],
        "history": [
            {
                "id": history.id,
                "action": history.action,
                "actor": history.actor,
                "from_status": history.from_status,
                "to_status": history.to_status,
                "test_question": decrypt_if_needed(history.test_question),
                "test_answer": decrypt_if_needed(history.test_answer),
                "test_source": history.test_source,
                "test_passed": history.test_passed,
                "created_at": history.created_at,
            }
            for history in histories
        ],
        "related_resolutions": [
            {
                "alert_id": related.id,
                "reason": related.reason,
                "test_question": decrypt_if_needed(related.test_question),
                "test_answer": decrypt_if_needed(related.test_answer),
                "test_source": related.test_source,
                "resolved_by": related.assigned_to,
                "resolved_at": related.resolved_at,
            }
            for related in related_alerts
        ],
    }


@router.post("/operations/alerts/{alert_id}/test")
async def test_operations_alert_answer(
    alert_id: int,
    body: OperationsAlertTestRequest,
    db: Session = Depends(get_db),
    current_user: str = Depends(verify_admin),
):
    alert = db.query(OperationsAlert).filter(OperationsAlert.id == alert_id).first()
    if not alert:
        raise HTTPException(status_code=404, detail="운영 알림을 찾을 수 없습니다.")
    question = body.question.strip()
    if not question:
        raise HTTPException(status_code=400, detail="테스트 질문을 입력해 주세요.")
    original_messages = db.query(ChatMessage).filter(
        ChatMessage.session_id == alert.session_id,
        ChatMessage.role.in_(("user", "assistant")),
    ).order_by(ChatMessage.created_at, ChatMessage.id).all()
    history = [
        HistoryMessage(role=message.role, content=decrypt_if_needed(message.content) or "")
        for message in original_messages[-8:]
    ]
    test_session_id = f"admin-test-{uuid4().hex}"
    try:
        response = await run_chat_preview(
            ChatRequest(session_id=test_session_id, message=question, history=history),
            db,
        )
        test_answer = response.answer
        test_source = response.source
    finally:
        db.rollback()
        db.query(OperationsAlert).filter(OperationsAlert.session_id == test_session_id).delete(synchronize_session=False)
        db.query(CancelRequest).filter(CancelRequest.session_id == test_session_id).delete(synchronize_session=False)
        db.query(ChatLog).filter(ChatLog.session_id == test_session_id).delete(synchronize_session=False)
        db.query(ChatMessage).filter(ChatMessage.session_id == test_session_id).delete(synchronize_session=False)
        db.query(ChatSession).filter(ChatSession.id == test_session_id).delete(synchronize_session=False)
        db.commit()

    alert = db.query(OperationsAlert).filter(OperationsAlert.id == alert_id).first()
    alert.test_question = maybe_encrypt(question)
    alert.test_answer = maybe_encrypt(test_answer)
    alert.test_source = test_source
    alert.test_passed = False
    alert.tested_by = current_user
    alert.tested_at = datetime.now()
    if alert.status == "open":
        previous_status = alert.status
        alert.status = "checking"
        alert.assigned_to = current_user
        _add_operations_alert_history(db, alert, "checking_started", current_user, previous_status)
    _add_operations_alert_history(db, alert, "answer_tested", current_user)
    db.commit()
    return {
        "question": question,
        "answer": test_answer,
        "source": test_source,
        "tested_by": current_user,
        "tested_at": alert.tested_at,
    }


@router.patch("/operations/alerts/{alert_id}")
def update_operations_alert(
    alert_id: int,
    body: OperationsAlertUpdate,
    db: Session = Depends(get_db),
    current_user: str = Depends(verify_admin),
):
    if body.status not in {"open", "checking", "resolved"}:
        raise HTTPException(status_code=400, detail="지원하지 않는 알림 상태입니다.")
    alert = db.query(OperationsAlert).filter(OperationsAlert.id == alert_id).first()
    if not alert:
        raise HTTPException(status_code=404, detail="운영 알림을 찾을 수 없습니다.")

    previous_status = alert.status
    if body.test_question is not None:
        next_question = body.test_question.strip()
        current_question = decrypt_if_needed(alert.test_question) or ""
        if next_question != current_question:
            alert.test_answer = None
            alert.test_source = None
            alert.test_passed = False
            alert.tested_by = None
            alert.tested_at = None
        alert.test_question = maybe_encrypt(next_question) if next_question else None

    for field_name in ("note",):
        value = getattr(body, field_name)
        if value is not None:
            stripped = value.strip()
            setattr(alert, field_name, maybe_encrypt(stripped) if stripped else None)
    if body.test_passed is not None:
        if body.test_passed and not alert.test_answer:
            raise HTTPException(status_code=400, detail="먼저 수정 후 답변 테스트를 실행해 주세요.")
        alert.test_passed = body.test_passed

    if body.status == "resolved":
        if not alert.tested_at or not alert.test_answer:
            raise HTTPException(status_code=400, detail="완료 전에 같은 질문으로 수정 후 답변 테스트를 실행해 주세요.")
        latest_test = db.query(OperationsAlertHistory).filter(
            OperationsAlertHistory.alert_id == alert.id,
            OperationsAlertHistory.action == "answer_tested",
        ).order_by(OperationsAlertHistory.id.desc()).first()
        latest_work = db.query(OperationsAlertHistory).filter(
            OperationsAlertHistory.alert_id == alert.id,
            OperationsAlertHistory.action.in_(("checking_started", "work_updated")),
        ).order_by(OperationsAlertHistory.id.desc()).first()
        if not latest_test or (latest_work and latest_test.id <= latest_work.id):
            raise HTTPException(status_code=400, detail="마지막 작업 내용 저장 이후의 답변 테스트가 필요합니다.")
        if not alert.test_passed:
            raise HTTPException(status_code=400, detail="테스트 답변이 원하는 결과인지 확인해 주세요.")

    alert.status = body.status
    alert.assigned_to = current_user if body.status in {"checking", "resolved"} else None
    alert.resolved_at = datetime.now() if body.status == "resolved" else None
    action = "resolved" if body.status == "resolved" else "checking_started" if previous_status != "checking" and body.status == "checking" else "work_updated"
    _add_operations_alert_history(db, alert, action, current_user, previous_status)
    db.commit()
    db.refresh(alert)
    create_audit_log(
        db,
        "operations_alert_updated",
        "operations_alert",
        str(alert.id),
        f"status={body.status}, action={action}",
        actor=current_user,
    )
    return _serialize_operations_alert(alert)


# ── 커스텀 데이터 관리 ──────────────────────────────────────────


class CreateTableRequest(BaseModel):
    name: str
    description: str = ""


class CreateColumnRequest(BaseModel):
    column_name: str
    column_type: str = "text"  # text | number | date


class UpsertRowRequest(BaseModel):
    data: dict


class UpdateColumnRequest(BaseModel):
    column_name: str


class ReorderColumnRequest(BaseModel):
    direction: str  # "up" | "down"


@router.get("/data-tables")
def list_data_tables(db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    tables = db.query(CustomTable).order_by(CustomTable.created_at.desc()).all()
    inspector = sa_inspect(db.bind)
    existing_tables = set(inspector.get_table_names())
    result = []
    for t in tables:
        real_table = f"cdata_{t.id}"
        if real_table in existing_tables:
            try:
                row_count = db.execute(text(f'SELECT COUNT(*) FROM "{real_table}"')).scalar() or 0  # noqa: S608
            except Exception:
                row_count = 0
        else:
            row_count = 0
        result.append({"id": t.id, "name": t.name, "description": t.description, "row_count": row_count, "created_at": t.created_at})
    return {"tables": result}


@router.post("/data-tables", status_code=201)
def create_data_table(body: CreateTableRequest, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    if not body.name.strip():
        raise HTTPException(status_code=400, detail="테이블 이름을 입력해주세요.")
    table = CustomTable(name=body.name.strip(), description=body.description.strip())
    db.add(table)
    db.commit()
    db.refresh(table)
    real_table = f"cdata_{table.id}"
    db.execute(text(f'CREATE TABLE IF NOT EXISTS "{real_table}" (id SERIAL PRIMARY KEY, created_at TIMESTAMP WITH TIME ZONE DEFAULT NOW(), updated_at TIMESTAMP WITH TIME ZONE DEFAULT NOW())'))  # noqa: S608
    db.commit()
    create_audit_log(db, "data_table_created", "custom_table", str(table.id), body.name)
    return {"id": table.id, "name": table.name, "description": table.description}


@router.get("/data-tables/export-all")
def export_all_data_tables(db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    tables = db.query(CustomTable).order_by(CustomTable.created_at.asc()).all()
    inspector = sa_inspect(db.bind)
    existing_tables = set(inspector.get_table_names())

    wb = Workbook()
    ws_index = wb.active
    ws_index.title = "개요"
    ws_index.append(["테이블명", "설명", "행 수", "생성일시"])

    used_sheet_names: set[str] = {"개요"}

    for t in tables:
        real_table = f"cdata_{t.id}"
        cols = db.query(CustomColumn).filter(CustomColumn.table_id == t.id).order_by(CustomColumn.sort_order, CustomColumn.id).all()
        col_names = [c.column_name for c in cols]

        rows: list = []
        if real_table in existing_tables and col_names:
            try:
                select_cols = ", ".join([_qi(cn) for cn in col_names])
                raw = db.execute(text(f'SELECT id, {select_cols}, created_at FROM "{real_table}" ORDER BY id')).fetchall()  # noqa: S608
                rows = list(raw)
            except Exception:
                rows = []

        ws_index.append([t.name, t.description or "", len(rows), t.created_at.isoformat() if t.created_at else ""])

        # 시트 이름 충돌 방지
        sheet_name = t.name[:28]
        if sheet_name in used_sheet_names:
            sheet_name = f"{sheet_name[:25]}_{t.id}"
        used_sheet_names.add(sheet_name)

        ws = wb.create_sheet(title=sheet_name)
        ws.append(["ID"] + col_names + ["생성일시"])
        for r in rows:
            row_vals = list(r[1:-1]) if col_names else []
            created = r[-1].isoformat() if r[-1] else ""
            ws.append([r[0]] + row_vals + [created])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"all_data_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.delete("/data-tables/{table_id}")
def delete_data_table(table_id: int, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    table = db.query(CustomTable).filter(CustomTable.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="테이블을 찾을 수 없습니다.")
    table_name = table.name  # commit 후엔 ORM 인스턴스가 만료되므로 미리 보관
    real_table = f"cdata_{table_id}"
    db.execute(text(f'DROP TABLE IF EXISTS "{real_table}"'))  # noqa: S608
    db.query(CustomColumn).filter(CustomColumn.table_id == table_id).delete()
    db.delete(table)
    db.commit()
    create_audit_log(db, "data_table_deleted", "custom_table", str(table_id), table_name)
    return {"message": "삭제되었습니다."}


@router.get("/data-tables/{table_id}")
def get_data_table(table_id: int, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    table = db.query(CustomTable).filter(CustomTable.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="테이블을 찾을 수 없습니다.")
    columns = db.query(CustomColumn).filter(CustomColumn.table_id == table_id).order_by(CustomColumn.sort_order, CustomColumn.id).all()
    col_names = [c.column_name for c in columns]
    real_table = f"cdata_{table_id}"
    rows: list[dict] = []
    inspector = sa_inspect(db.bind)
    if real_table in inspector.get_table_names():
        select_cols = ", ".join([_qi(cn) for cn in col_names]) if col_names else "1 as _empty"
        raw_rows = db.execute(text(f'SELECT id, {select_cols}, created_at FROM "{real_table}" ORDER BY id DESC')).fetchall()  # noqa: S608
        for r in raw_rows:
            row_data = dict(zip(col_names, list(r)[1:-1]))
            rows.append({"id": r[0], "data": row_data, "created_at": r[-1]})
    return {
        "id": table.id,
        "name": table.name,
        "description": table.description,
        "columns": [{"id": c.id, "column_name": c.column_name, "column_type": c.column_type, "sort_order": c.sort_order} for c in columns],
        "rows": rows,
    }


_COL_TYPE_MAP = {"text": "TEXT", "number": "NUMERIC", "date": "DATE"}


def _qi(name: str) -> str:
    """SQL 식별자 안전 인용: 내부 큰따옴표를 이스케이프해 인젝션을 차단한다. (한글 등 유니코드 컬럼명은 그대로 허용)"""
    return '"' + str(name).replace('"', '""') + '"'


def _validate_col_name(name: str) -> str:
    """사용자 입력 컬럼명 검증: 위험 문자(큰따옴표/제어문자) 차단 + 길이 제한. 한글·공백은 허용."""
    cleaned = (name or "").strip()
    if not cleaned:
        raise HTTPException(status_code=400, detail="컬럼 이름을 입력해주세요.")
    if len(cleaned) > 63 or '"' in cleaned or any(ord(ch) < 32 for ch in cleaned):
        raise HTTPException(status_code=400, detail='컬럼 이름에 큰따옴표(")나 제어문자는 사용할 수 없으며 63자 이하여야 합니다.')
    return cleaned


def _validate_row_keys(db: Session, table_id: int, data: dict) -> None:
    """행 데이터의 키가 실제 등록된 컬럼인지 검증 (식별자 인젝션·오타 키 차단)."""
    valid = {c.column_name for c in db.query(CustomColumn).filter(CustomColumn.table_id == table_id).all()}
    invalid = [k for k in data if k not in valid]
    if invalid:
        raise HTTPException(status_code=400, detail=f"존재하지 않는 컬럼입니다: {', '.join(invalid)}")


@router.post("/data-tables/{table_id}/columns", status_code=201)
def add_column(table_id: int, body: CreateColumnRequest, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    if not db.query(CustomTable).filter(CustomTable.id == table_id).first():
        raise HTTPException(status_code=404, detail="테이블을 찾을 수 없습니다.")
    if not body.column_name.strip():
        raise HTTPException(status_code=400, detail="컬럼 이름을 입력해주세요.")
    if body.column_type not in ("text", "number", "date"):
        raise HTTPException(status_code=400, detail="컬럼 타입은 text, number, date 중 하나여야 합니다.")
    max_order = db.query(CustomColumn).filter(CustomColumn.table_id == table_id).count()
    col_name = _validate_col_name(body.column_name)
    col = CustomColumn(table_id=table_id, column_name=col_name, column_type=body.column_type, sort_order=max_order)
    db.add(col)
    db.commit()
    db.refresh(col)
    sql_type = _COL_TYPE_MAP.get(body.column_type, "TEXT")
    real_table = f"cdata_{table_id}"
    db.execute(text(f'ALTER TABLE "{real_table}" ADD COLUMN IF NOT EXISTS {_qi(col_name)} {sql_type}'))  # noqa: S608
    db.commit()
    return {"id": col.id, "column_name": col.column_name, "column_type": col.column_type, "sort_order": col.sort_order}


@router.delete("/data-tables/{table_id}/columns/{column_id}")
def delete_column(table_id: int, column_id: int, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    col = db.query(CustomColumn).filter(CustomColumn.id == column_id, CustomColumn.table_id == table_id).first()
    if not col:
        raise HTTPException(status_code=404, detail="컬럼을 찾을 수 없습니다.")
    col_name = col.column_name
    db.delete(col)
    db.commit()
    real_table = f"cdata_{table_id}"
    db.execute(text(f'ALTER TABLE "{real_table}" DROP COLUMN IF EXISTS {_qi(col_name)}'))  # noqa: S608
    db.commit()
    return {"message": "컬럼이 삭제되었습니다."}


@router.put("/data-tables/{table_id}/columns/{column_id}")
def rename_column(table_id: int, column_id: int, body: UpdateColumnRequest, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    col = db.query(CustomColumn).filter(CustomColumn.id == column_id, CustomColumn.table_id == table_id).first()
    if not col:
        raise HTTPException(status_code=404, detail="컬럼을 찾을 수 없습니다.")
    new_name = _validate_col_name(body.column_name)
    old_name = col.column_name
    real_table = f"cdata_{table_id}"
    db.execute(text(f'ALTER TABLE "{real_table}" RENAME COLUMN {_qi(old_name)} TO {_qi(new_name)}'))  # noqa: S608
    col.column_name = new_name
    db.commit()
    return {"id": col.id, "column_name": col.column_name, "column_type": col.column_type, "sort_order": col.sort_order}


@router.post("/data-tables/{table_id}/columns/{column_id}/reorder")
def reorder_column(table_id: int, column_id: int, body: ReorderColumnRequest, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    if body.direction not in {"up", "down"}:
        raise HTTPException(status_code=400, detail="direction은 up 또는 down이어야 합니다.")
    columns = db.query(CustomColumn).filter(CustomColumn.table_id == table_id).order_by(CustomColumn.sort_order, CustomColumn.id).all()
    idx = next((i for i, c in enumerate(columns) if c.id == column_id), None)
    if idx is None:
        raise HTTPException(status_code=404, detail="컬럼을 찾을 수 없습니다.")
    swap_idx = idx - 1 if body.direction == "up" else idx + 1
    if swap_idx < 0 or swap_idx >= len(columns):
        return {"message": "이동할 수 없습니다."}
    columns[idx], columns[swap_idx] = columns[swap_idx], columns[idx]
    for i, col in enumerate(columns):
        col.sort_order = i
    db.commit()
    return {"message": "컬럼 순서를 변경했습니다."}


@router.post("/data-tables/{table_id}/rows", status_code=201)
def add_row(table_id: int, body: UpsertRowRequest, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    if not db.query(CustomTable).filter(CustomTable.id == table_id).first():
        raise HTTPException(status_code=404, detail="테이블을 찾을 수 없습니다.")
    real_table = f"cdata_{table_id}"
    if not body.data:
        row_id = db.execute(text(f'INSERT INTO "{real_table}" DEFAULT VALUES RETURNING id, created_at')).fetchone()  # noqa: S608
    else:
        _validate_row_keys(db, table_id, body.data)
        col_sql = ", ".join([_qi(k) for k in body.data])
        val_sql = ", ".join([f":v{i}" for i in range(len(body.data))])
        params = {f"v{i}": v for i, v in enumerate(body.data.values())}
        row_id = db.execute(text(f'INSERT INTO "{real_table}" ({col_sql}) VALUES ({val_sql}) RETURNING id, created_at'), params).fetchone()  # noqa: S608
    db.commit()
    return {"id": row_id[0], "data": body.data, "created_at": row_id[1]}


@router.put("/data-tables/{table_id}/rows/{row_id}")
def update_row(table_id: int, row_id: int, body: UpsertRowRequest, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    if not db.query(CustomTable).filter(CustomTable.id == table_id).first():
        raise HTTPException(status_code=404, detail="테이블을 찾을 수 없습니다.")
    real_table = f"cdata_{table_id}"
    if not body.data:
        db.execute(text(f'UPDATE "{real_table}" SET updated_at = NOW() WHERE id = :id'), {"id": row_id})  # noqa: S608
    else:
        _validate_row_keys(db, table_id, body.data)
        set_sql = ", ".join([f'{_qi(k)} = :v{i}' for i, k in enumerate(body.data)])
        params = {f"v{i}": v for i, v in enumerate(body.data.values())}
        params["id"] = row_id
        db.execute(text(f'UPDATE "{real_table}" SET {set_sql}, updated_at = NOW() WHERE id = :id'), params)  # noqa: S608
    db.commit()
    return {"id": row_id, "data": body.data}


@router.delete("/data-tables/{table_id}/rows/{row_id}")
def delete_row(table_id: int, row_id: int, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    real_table = f"cdata_{table_id}"
    db.execute(text(f'DELETE FROM "{real_table}" WHERE id = :id'), {"id": row_id})  # noqa: S608
    db.commit()
    return {"message": "행이 삭제되었습니다."}


@router.get("/data-tables/{table_id}/export")
def export_data_table(table_id: int, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    table = db.query(CustomTable).filter(CustomTable.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="테이블을 찾을 수 없습니다.")
    columns = db.query(CustomColumn).filter(CustomColumn.table_id == table_id).order_by(CustomColumn.sort_order, CustomColumn.id).all()
    col_names = [c.column_name for c in columns]
    real_table = f"cdata_{table_id}"
    if col_names:
        select_cols = ", ".join([_qi(cn) for cn in col_names])
        raw_rows = db.execute(text(f'SELECT id, {select_cols}, created_at FROM "{real_table}" ORDER BY id')).fetchall()  # noqa: S608
    else:
        raw_rows = db.execute(text(f'SELECT id, created_at FROM "{real_table}" ORDER BY id')).fetchall()  # noqa: S608
    wb = Workbook()
    ws = wb.active
    ws.title = table.name[:31]
    ws.append(["ID"] + col_names + ["생성일시"])
    for r in raw_rows:
        row_values = list(r[1:-1]) if col_names else []
        ws.append([r[0]] + row_values + [r[-1].isoformat() if r[-1] else ""])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{table.name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/data-tables/{table_id}/import")
async def import_table_data(table_id: int, file: UploadFile = File(...), db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    table = db.query(CustomTable).filter(CustomTable.id == table_id).first()
    if not table:
        raise HTTPException(status_code=404, detail="테이블을 찾을 수 없습니다.")

    fname = (file.filename or "").lower()
    if not (fname.endswith(".csv") or fname.endswith(".xlsx") or fname.endswith(".xls")):
        raise HTTPException(status_code=400, detail="CSV 또는 Excel(.xlsx/.xls) 파일만 업로드할 수 있습니다.")

    cols = db.query(CustomColumn).filter(CustomColumn.table_id == table_id).order_by(CustomColumn.sort_order, CustomColumn.id).all()
    col_names = [c.column_name for c in cols]
    if not col_names:
        raise HTTPException(status_code=400, detail="컬럼을 먼저 추가해주세요.")

    content = await file.read()

    if fname.endswith(".csv"):
        text_content = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text_content))
        raw_rows: list[dict] = [dict(r) for r in reader]
    else:
        from openpyxl import load_workbook as _load_wb
        wb_in = _load_wb(io.BytesIO(content), read_only=True)
        ws_in = wb_in.active
        headers = [str(cell.value) if cell.value is not None else "" for cell in next(ws_in.iter_rows(max_row=1))]
        raw_rows = []
        for row in ws_in.iter_rows(min_row=2, values_only=True):
            raw_rows.append(dict(zip(headers, [str(v) if v is not None else "" for v in row])))
        wb_in.close()

    real_table = f"cdata_{table_id}"
    count = 0
    for raw in raw_rows:
        data = {cn: str(raw[cn]) for cn in col_names if cn in raw and raw[cn] not in (None, "")}
        if not data:
            continue
        col_sql = ", ".join([_qi(k) for k in data])
        val_sql = ", ".join([f":v{i}" for i in range(len(data))])
        params = {f"v{i}": v for i, v in enumerate(data.values())}
        db.execute(text(f'INSERT INTO "{real_table}" ({col_sql}) VALUES ({val_sql})'), params)  # noqa: S608
        count += 1

    db.commit()
    return {"message": f"{count}개 행을 가져왔습니다.", "count": count}


# ── DB 브라우저 ──────────────────────────────────────────────


@router.get("/db/tables")
def list_db_tables(db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    inspector = sa_inspect(db.bind)
    tables = sorted(inspector.get_table_names())
    custom_meta: dict[str, tuple[str, str]] = {}
    for ct in db.query(CustomTable).all():
        custom_meta[f"cdata_{ct.id}"] = (ct.name, ct.description or "")
    result = []
    for table_name in tables:
        try:
            count = db.execute(text(f'SELECT COUNT(*) FROM "{table_name}"')).scalar() or 0  # noqa: S608
        except Exception:
            count = -1
        columns = [col["name"] for col in inspector.get_columns(table_name)]
        if table_name in custom_meta:
            display_name, description = custom_meta[table_name]
            display_name = f"[데이터] {display_name}"
        else:
            display_name = table_name
            description = TABLE_DESCRIPTIONS.get(table_name, "")
        result.append({"name": table_name, "display_name": display_name, "description": description, "row_count": count, "columns": columns})
    return {"tables": result}


@router.get("/db/tables/{table_name}")
def browse_db_table(
    table_name: str,
    page: int = Query(1, ge=1),
    limit: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    _: None = Depends(verify_admin),
):
    inspector = sa_inspect(db.bind)
    valid_tables = inspector.get_table_names()
    if table_name not in valid_tables:
        raise HTTPException(status_code=404, detail="테이블을 찾을 수 없습니다.")

    columns = [col["name"] for col in inspector.get_columns(table_name)]
    offset = (page - 1) * limit
    total = db.execute(text(f'SELECT COUNT(*) FROM "{table_name}"')).scalar() or 0  # noqa: S608
    try:
        rows_result = db.execute(
            text(f'SELECT * FROM "{table_name}" ORDER BY id DESC LIMIT :limit OFFSET :offset'),  # noqa: S608
            {"limit": limit, "offset": offset},
        ).fetchall()
    except Exception:
        rows_result = db.execute(
            text(f'SELECT * FROM "{table_name}" LIMIT :limit OFFSET :offset'),  # noqa: S608
            {"limit": limit, "offset": offset},
        ).fetchall()

    def _serialize(v):
        if isinstance(v, datetime):
            return v.isoformat()
        if isinstance(v, (date, time)):
            return str(v)
        if isinstance(v, str):
            return decrypt_if_needed(v) or v
        return v

    rows = [dict(zip(columns, [_serialize(v) for v in row])) for row in rows_result]
    return {
        "columns": columns,
        "rows": rows,
        "total": total,
        "page": page,
        "limit": limit,
        "editable": table_name in EDITABLE_TABLES,
        "droppable": _is_droppable(table_name),
        "restriction_reason": _restriction_reason(table_name),
        "protected_columns": sorted(PROTECTED_COLUMNS),
    }


# ─────────────────────────────────────────────────────────────────────────
# DB 브라우저 행 편집·삭제 (안전한 4개 테이블만 화이트리스트)
# ─────────────────────────────────────────────────────────────────────────

EDITABLE_TABLES = {"faqs", "chat_logs", "processing_logs", "cancel_requests"}

# 저장 시 카테고리 토글 상태에 따라 enc:: 자동 처리되는 컬럼
ENCRYPT_AWARE_COLUMNS: dict[str, set[str]] = {
    "faqs": {"category", "question", "answer", "keywords_json", "aliases_json", "search_hints_json", "source_files_json"},
    "chat_logs": {"question", "answer", "retrieval_chunks", "error"},
    # processing_logs / cancel_requests는 평문 저장
}

# 수정·삭제 시 보호되는 컬럼 (편집 불가)
PROTECTED_COLUMNS = {"id", "created_at", "updated_at"}

# 테이블 자체를 DROP 가능한 화이트리스트 (시스템 영향 없음)
DROPPABLE_TABLES = {"chat_logs", "processing_logs", "cancel_requests", "admin_audit_logs"}
# cdata_* 동적 사용자 정의 테이블은 prefix 매칭으로 별도 허용


class DbRowUpdate(BaseModel):
    values: dict[str, object]


@router.put("/db/tables/{table_name}/rows/{row_id}")
def update_db_row(
    table_name: str,
    row_id: int,
    body: DbRowUpdate,
    db: Session = Depends(get_db),
    _: str = Depends(verify_admin),
):
    if table_name not in EDITABLE_TABLES:
        raise HTTPException(status_code=403, detail="이 테이블은 편집할 수 없습니다.")

    inspector = sa_inspect(db.bind)
    if table_name not in inspector.get_table_names():
        raise HTTPException(status_code=404, detail="테이블을 찾을 수 없습니다.")

    valid_columns = {col["name"] for col in inspector.get_columns(table_name)}
    enc_columns = ENCRYPT_AWARE_COLUMNS.get(table_name, set())

    updates: dict[str, object] = {}
    for col, raw_value in body.values.items():
        if col in PROTECTED_COLUMNS or col not in valid_columns:
            continue
        value = raw_value
        if isinstance(value, str) and col in enc_columns:
            value = maybe_encrypt(value)
        updates[col] = value

    if not updates:
        raise HTTPException(status_code=400, detail="수정할 컬럼이 없습니다.")

    set_clause = ", ".join([f'"{k}" = :{k}' for k in updates])
    params = {**updates, "row_id": row_id}
    result = db.execute(
        text(f'UPDATE "{table_name}" SET {set_clause} WHERE id = :row_id'),  # noqa: S608
        params,
    )
    if result.rowcount == 0:
        raise HTTPException(status_code=404, detail="행을 찾을 수 없습니다.")
    db.commit()

    create_audit_log(
        db,
        "db_row_update",
        table_name,
        str(row_id),
        f"수정된 컬럼: {', '.join(updates.keys())}",
    )
    return {"message": "수정되었습니다.", "table": table_name, "row_id": row_id, "updated": list(updates.keys())}


@router.delete("/db/tables/{table_name}/rows/{row_id}")
def delete_db_row(
    table_name: str,
    row_id: int,
    db: Session = Depends(get_db),
    _: str = Depends(verify_admin),
):
    if table_name not in EDITABLE_TABLES:
        raise HTTPException(status_code=403, detail="이 테이블은 삭제할 수 없습니다.")

    inspector = sa_inspect(db.bind)
    if table_name not in inspector.get_table_names():
        raise HTTPException(status_code=404, detail="테이블을 찾을 수 없습니다.")

    try:
        result = db.execute(
            text(f'DELETE FROM "{table_name}" WHERE id = :row_id'),  # noqa: S608
            {"row_id": row_id},
        )
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"삭제 실패: {exc}") from exc

    if result.rowcount == 0:
        raise HTTPException(status_code=404, detail="행을 찾을 수 없습니다.")
    db.commit()

    create_audit_log(db, "db_row_delete", table_name, str(row_id), "행 삭제")
    return {"message": "삭제되었습니다.", "table": table_name, "row_id": row_id}


def _is_droppable(table_name: str) -> bool:
    if table_name in DROPPABLE_TABLES:
        return True
    if table_name.startswith("cdata_"):
        return True
    return False


def _restriction_reason(table_name: str) -> str | None:
    """편집·삭제 불가 사유를 사람이 읽을 수 있는 문구로 반환. 가능한 테이블이면 None."""
    if table_name in EDITABLE_TABLES or _is_droppable(table_name):
        return None
    reasons: dict[str, str] = {
        "chunks": "RAG 검색 인덱스(FAISS)와 1:1로 묶여 있어 직접 수정하면 검색이 깨집니다. 문서 검토 탭에서 재인덱싱으로만 변경하세요.",
        "documents": "원본 문서 메타. 문서 검토 탭의 승인·반려·삭제 흐름으로만 관리됩니다.",
        "chat_messages": "사용자 대화 본체. 수정·삭제하면 대화 이력이 깨집니다.",
        "chat_sessions": "세션 식별자. 변경 시 모든 메시지·로그가 끊깁니다.",
        "admin_users": "관리자 권한 목록. 권한 관리 탭에서만 안전하게 수정하세요.",
        "prompt_configs": "시스템 프롬프트. 프롬프트 탭에서 안전하게 편집하세요.",
        "faqs": "FAQ 콘텐츠. FAQ 관리 탭에서 안전하게 편집하세요.",
        "custom_tables": "사용자 정의 테이블 메타. 데이터 관리 탭에서 관리됩니다.",
        "custom_columns": "사용자 정의 컬럼 정의. 데이터 관리 탭에서 관리됩니다.",
    }
    return reasons.get(table_name, "시스템 무결성 보호를 위해 직접 편집·삭제가 차단되어 있습니다.")


@router.delete("/db/tables/{table_name}")
def drop_db_table(
    table_name: str,
    db: Session = Depends(get_db),
    _: str = Depends(verify_admin),
):
    if not _is_droppable(table_name):
        raise HTTPException(status_code=403, detail="이 테이블은 삭제할 수 없습니다.")

    inspector = sa_inspect(db.bind)
    if table_name not in inspector.get_table_names():
        raise HTTPException(status_code=404, detail="테이블을 찾을 수 없습니다.")

    # cdata_* 는 custom_tables 메타도 같이 정리
    if table_name.startswith("cdata_"):
        try:
            cdata_id = int(table_name[len("cdata_"):])
            db.execute(text("DELETE FROM custom_columns WHERE table_id = :tid"), {"tid": cdata_id})
            db.execute(text("DELETE FROM custom_tables WHERE id = :tid"), {"tid": cdata_id})
        except (ValueError, Exception):
            pass

    try:
        db.execute(text(f'DROP TABLE IF EXISTS "{table_name}" CASCADE'))  # noqa: S608
        db.commit()
    except Exception as exc:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"테이블 삭제 실패: {exc}") from exc

    create_audit_log(db, "db_table_drop", table_name, table_name, "테이블 DROP")
    return {"message": f"{table_name} 테이블을 삭제했습니다.", "table": table_name}


def _is_chat_model(model_id: str) -> bool:
    if ":" in model_id:
        return False
    EXCLUDED = ("instruct", "realtime", "audio", "tts", "whisper", "dall", "embedding", "moderation", "vision")
    if any(kw in model_id for kw in EXCLUDED):
        return False
    prefixes = ("gpt-5", "gpt-4", "gpt-3.5-turbo", "o1", "o3", "o4", "chatgpt")
    return any(model_id.startswith(p) for p in prefixes)


@router.get("/settings/model")
async def get_model_settings(_: None = Depends(verify_admin)):
    settings = get_settings()
    try:
        openai_client = AsyncOpenAI(api_key=settings.openai_api_key)
        response = await openai_client.models.list()
        chat_models = sorted(
            [m for m in response.data if _is_chat_model(m.id)],
            key=lambda m: m.created,
            reverse=True,
        )
        available_models = [m.id for m in chat_models]
    except Exception:
        available_models = ["gpt-4o", "gpt-4o-mini", "gpt-4-turbo", "gpt-3.5-turbo"]
    return {"current_model": get_active_model(), "available_models": available_models}


@router.put("/settings/model")
def change_model(body: ModelChangeRequest, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    # 모델 설정은 DB(app_settings)에 저장한다. (.env 방식은 배포 재생성·환경변수 우선·멀티워커로 인해 유지 안 됨)
    model_name = (body.model_name or "").strip()
    if not _is_chat_model(model_name):
        raise HTTPException(status_code=400, detail="유효한 채팅 모델 ID가 아닙니다.")
    set_active_model(db, model_name)
    create_audit_log(db, "model_changed", "system", "model_name", model_name)
    return {"message": f"모델을 {model_name}으로 변경했습니다.", "model_name": model_name}


@router.get("/security-vault/status")
def get_security_vault_status(
    db: Session = Depends(get_db),
    current_user: str = Depends(verify_superadmin),
):
    configured = _vault_password_setting(db) is not None
    return {
        "configured": configured,
        "can_setup": not configured and current_user == get_settings().admin_email,
        "unlock_minutes": VAULT_TOKEN_MINUTES,
        "protected_keys": ["ENCRYPTION_KEY", "JWT_SECRET", "ADMIN_PASSWORD"],
    }


@router.post("/security-vault/setup")
def setup_security_vault(
    body: VaultPasswordRequest,
    db: Session = Depends(get_db),
    current_user: str = Depends(verify_superadmin),
):
    if _vault_password_setting(db) is not None:
        raise HTTPException(status_code=409, detail="보안 정보 보관 비밀번호가 이미 설정되어 있습니다.")
    _validate_vault_password(body.password)
    db.add(AppSetting(key=VAULT_PASSWORD_SETTING_KEY, value=_hash_vault_password(body.password)))
    _ensure_vault_defaults(db)
    db.commit()
    create_audit_log(db, "security_vault_configured", "security_vault", "setup", actor=current_user)
    return {
        "message": "보안 정보 보관 비밀번호를 설정했습니다.",
        "vault_token": _create_vault_token(current_user, db),
        "expires_in_seconds": VAULT_TOKEN_MINUTES * 60,
    }


@router.post("/security-vault/unlock")
def unlock_security_vault(
    body: VaultPasswordRequest,
    db: Session = Depends(get_db),
    current_user: str = Depends(verify_superadmin),
):
    setting = _vault_password_setting(db)
    if setting is None:
        raise HTTPException(status_code=409, detail="먼저 보안 정보 보관 비밀번호를 설정해 주세요.")
    if not _verify_vault_password(body.password, setting.value or ""):
        create_audit_log(db, "security_vault_unlock_failed", "security_vault", "unlock", actor=current_user)
        raise HTTPException(status_code=403, detail="보관 비밀번호가 올바르지 않습니다.")
    create_audit_log(db, "security_vault_unlocked", "security_vault", "unlock", actor=current_user)
    return {
        "message": "보안 정보 잠금을 해제했습니다.",
        "vault_token": _create_vault_token(current_user, db),
        "expires_in_seconds": VAULT_TOKEN_MINUTES * 60,
    }


@router.post("/security-vault/extend")
def extend_security_vault(
    x_vault_token: str | None = Header(None, alias="X-Vault-Token"),
    db: Session = Depends(get_db),
    current_user: str = Depends(verify_superadmin),
):
    _require_vault_token(current_user, x_vault_token, db)
    create_audit_log(db, "security_vault_extended", "security_vault", "extend", actor=current_user)
    return {
        "message": "보안 정보 자동 잠금 시간을 연장했습니다.",
        "vault_token": _create_vault_token(current_user, db),
        "expires_in_seconds": VAULT_TOKEN_MINUTES * 60,
    }


@router.put("/security-vault/password")
def reset_security_vault_password(
    body: VaultPasswordRequest,
    x_vault_token: str | None = Header(None, alias="X-Vault-Token"),
    db: Session = Depends(get_db),
    current_user: str = Depends(verify_superadmin),
):
    _require_vault_token(current_user, x_vault_token, db)
    _validate_vault_password(body.password)
    setting = _vault_password_setting(db)
    if setting is None:
        raise HTTPException(status_code=409, detail="먼저 보안 정보 보관 비밀번호를 설정해 주세요.")
    setting.value = _hash_vault_password(body.password)
    db.commit()
    create_audit_log(db, "security_vault_password_reset", "security_vault", "password", actor=current_user)
    return {
        "message": "보안 정보 보관 비밀번호를 재설정했습니다.",
        "vault_token": _create_vault_token(current_user, db),
        "expires_in_seconds": VAULT_TOKEN_MINUTES * 60,
    }


@router.get("/security-vault/data")
def get_security_vault_data(
    x_vault_token: str | None = Header(None, alias="X-Vault-Token"),
    db: Session = Depends(get_db),
    current_user: str = Depends(verify_superadmin),
):
    _require_vault_token(current_user, x_vault_token, db)
    records = _ensure_vault_defaults(db)
    db.commit()
    create_audit_log(
        db,
        "security_vault_viewed",
        "security_vault",
        "credentials_and_environment",
        "허용된 보안정보와 환경설정 열람",
        actor=current_user,
    )
    return {
        "credentials": [_serialize_vault_secret(row) for row in records],
        "environment": _vault_environment_items(),
        "expires_in_seconds": VAULT_TOKEN_MINUTES * 60,
    }


@router.put("/security-vault/items/{secret_key}")
def save_security_vault_item(
    secret_key: str,
    body: VaultSecretPayload,
    x_vault_token: str | None = Header(None, alias="X-Vault-Token"),
    db: Session = Depends(get_db),
    current_user: str = Depends(verify_superadmin),
):
    _require_vault_token(current_user, x_vault_token, db)
    if secret_key not in VAULT_DEFAULTS:
        raise HTTPException(status_code=404, detail="지원하지 않는 보안 정보 항목입니다.")
    label = body.label.strip()
    if not label:
        raise HTTPException(status_code=400, detail="항목 이름을 입력해 주세요.")
    login_url = (body.login_url or "").strip()
    if login_url and not re.match(r"^https://", login_url, flags=re.IGNORECASE):
        raise HTTPException(status_code=400, detail="접속 주소는 https:// 형식이어야 합니다.")

    row = db.query(AdminSecretRecord).filter(AdminSecretRecord.secret_key == secret_key).first()
    if row is None:
        row = AdminSecretRecord(secret_key=secret_key, label=label)
        db.add(row)
    row.label = label
    row.login_url = login_url or None
    row.account_identifier = (body.account_identifier or "").strip() or None
    instance_identifier = (body.instance_identifier or "").strip()
    if instance_identifier and not re.fullmatch(r"i-[0-9a-fA-F]{8,17}", instance_identifier):
        raise HTTPException(status_code=400, detail="EC2 인스턴스 ID는 i-로 시작하는 올바른 형식이어야 합니다.")
    row.instance_identifier = instance_identifier or None
    if body.username is not None:
        row.username_encrypted = maybe_encrypt(body.username.strip()) or None
    if body.password is not None:
        row.password_encrypted = maybe_encrypt(body.password) or None
    if body.note is not None:
        row.note_encrypted = maybe_encrypt(body.note.strip()) or None
    row.updated_by = current_user
    db.commit()
    db.refresh(row)
    create_audit_log(
        db,
        "security_vault_item_saved",
        "security_vault",
        secret_key,
        "보안 정보 항목 수정",
        actor=current_user,
    )
    return _serialize_vault_secret(row)


@router.put("/password")
def change_password(body: PasswordChangeRequest, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    if not body.new_password or len(body.new_password) < 4:
        raise HTTPException(status_code=400, detail="비밀번호는 4자 이상이어야 합니다.")
    lines = ENV_PATH.read_text(encoding="utf-8").splitlines() if ENV_PATH.exists() else []
    updated = [line for line in lines if not line.startswith("ADMIN_PASSWORD=")]
    updated.append(f"ADMIN_PASSWORD={body.new_password}")
    ENV_PATH.write_text("\n".join(updated) + "\n", encoding="utf-8")
    get_settings.cache_clear()
    create_audit_log(db, "password_changed", "system", "admin_password")
    return {"message": "비밀번호를 변경했습니다."}


# ── 권한 관리 ──────────────────────────────────────────────────


class ChangeSuperadminRequest(BaseModel):
    new_email: str


@router.put("/settings/superadmin")
def change_superadmin(
    body: ChangeSuperadminRequest,
    db: Session = Depends(get_db),
    current_user: str = Depends(verify_superadmin),
):
    new_email = body.new_email.strip().lower()
    if not new_email or "@" not in new_email:
        raise HTTPException(status_code=400, detail="유효한 이메일을 입력해주세요.")
    if new_email == get_settings().admin_email:
        raise HTTPException(status_code=400, detail="현재 최상위 관리자 이메일과 동일합니다.")
    lines = ENV_PATH.read_text(encoding="utf-8").splitlines() if ENV_PATH.exists() else []
    updated = [line for line in lines if not line.startswith("ADMIN_EMAIL=")]
    updated.append(f"ADMIN_EMAIL={new_email}")
    ENV_PATH.write_text("\n".join(updated) + "\n", encoding="utf-8")
    get_settings.cache_clear()
    create_audit_log(db, "superadmin_changed", "system", "admin_email", current_user)
    return {"message": f"최상위 관리자를 {new_email}로 변경했습니다. 다시 로그인해주세요."}


class AddPermissionRequest(BaseModel):
    email: str


@router.get("/permissions/access")
def get_permission_access(current_user: str = Depends(verify_admin)):
    return {
        "current_user": current_user,
        "is_superadmin": current_user == get_settings().admin_email,
    }


@router.get("/permissions")
def list_permissions(db: Session = Depends(get_db), current_user: str = Depends(verify_superadmin)):
    users = db.query(AdminUser).order_by(AdminUser.created_at).all()
    return {
        "superadmin": get_settings().admin_email,
        "current_user": current_user,
        "admins": [
            {
                "email": u.email,
                "added_by": u.added_by,
                "created_at": u.created_at.isoformat() if u.created_at else None,
            }
            for u in users
        ],
    }


@router.post("/permissions", status_code=201)
def add_permission(body: AddPermissionRequest, db: Session = Depends(get_db), current_user: str = Depends(verify_superadmin)):
    if not body.email or "@" not in body.email:
        raise HTTPException(status_code=400, detail="유효한 이메일을 입력해주세요.")
    if db.query(AdminUser).filter(AdminUser.email == body.email).first():
        raise HTTPException(status_code=409, detail="이미 등록된 이메일입니다.")
    user = AdminUser(email=body.email, added_by=current_user)
    db.add(user)
    db.commit()
    create_audit_log(db, "permission_added", "admin_user", body.email, current_user)
    return {"message": f"{body.email}에 권한을 부여했습니다."}


@router.delete("/permissions/{email}")
def remove_permission(email: str, db: Session = Depends(get_db), current_user: str = Depends(verify_superadmin)):
    if email == get_settings().admin_email:
        raise HTTPException(status_code=400, detail="기본 관리자 이메일은 제거할 수 없습니다.")
    if email == current_user:
        raise HTTPException(status_code=400, detail="본인 계정은 제거할 수 없습니다. (자가 잠금 방지)")
    user = db.query(AdminUser).filter(AdminUser.email == email).first()
    if not user:
        raise HTTPException(status_code=404, detail="등록되지 않은 이메일입니다.")
    db.delete(user)
    db.commit()
    create_audit_log(db, "permission_removed", "admin_user", email, current_user)
    return {"message": f"{email}의 권한을 제거했습니다."}


# ── 암호화 설정 ─────────────────────────────────────────────────


class EncryptionToggleRequest(BaseModel):
    encrypt_enabled: bool


class EncryptionMigrateRequest(BaseModel):
    category: str
    direction: str  # "encrypt" | "decrypt"


def _count_encrypted(values: list[str | None]) -> int:
    return sum(1 for v in values if v and v.startswith(ENCRYPTED_PREFIX))


@router.get("/settings/encryption")
def get_encryption_settings(db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    settings = get_settings()

    faq_rows = db.query(FaqRecord).filter(FaqRecord.is_active.is_(True)).all()
    faq_enc = _count_encrypted([r.answer for r in faq_rows])

    prompt_rows = db.query(PromptConfig).all()
    prompt_enc = _count_encrypted([r.content for r in prompt_rows])

    doc_rows = db.query(DocumentRecord).filter(DocumentRecord.is_deleted.is_(False)).all()
    doc_enc = _count_encrypted([r.original_filename for r in doc_rows])

    return {
        "categories": [
            {
                "key": "faq",
                "label": "FAQ 내용",
                "encrypt_enabled": settings.encrypt_faq,
                "encrypted_count": faq_enc,
                "plain_count": len(faq_rows) - faq_enc,
                "total": len(faq_rows),
            },
            {
                "key": "prompt",
                "label": "프롬프트 내용",
                "encrypt_enabled": settings.encrypt_prompt,
                "encrypted_count": prompt_enc,
                "plain_count": len(prompt_rows) - prompt_enc,
                "total": len(prompt_rows),
            },
            {
                "key": "document",
                "label": "문서 파일명·검토내용",
                "encrypt_enabled": settings.encrypt_document,
                "encrypted_count": doc_enc,
                "plain_count": len(doc_rows) - doc_enc,
                "total": len(doc_rows),
            },
        ]
    }


@router.put("/settings/encryption/{category}")
def toggle_encryption(category: str, body: EncryptionToggleRequest, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    if category not in {"faq", "prompt", "document"}:
        raise HTTPException(status_code=400, detail="유효하지 않은 카테고리입니다.")
    env_key = f"ENCRYPT_{category.upper()}"
    env_value = "true" if body.encrypt_enabled else "false"
    lines = ENV_PATH.read_text(encoding="utf-8").splitlines() if ENV_PATH.exists() else []
    updated = [line for line in lines if not line.startswith(f"{env_key}=")]
    updated.append(f"{env_key}={env_value}")
    ENV_PATH.write_text("\n".join(updated) + "\n", encoding="utf-8")
    get_settings.cache_clear()
    create_audit_log(db, "encryption_toggled", "system", category, f"{env_key}={env_value}")
    label = "활성화" if body.encrypt_enabled else "비활성화"
    return {"message": f"{category} 암호화가 {label}되었습니다.", "category": category, "encrypt_enabled": body.encrypt_enabled}


@router.post("/settings/encryption/migrate")
def migrate_encryption(body: EncryptionMigrateRequest, db: Session = Depends(get_db), _: None = Depends(verify_admin)):
    if body.category not in {"faq", "prompt", "document"}:
        raise HTTPException(status_code=400, detail="유효하지 않은 카테고리입니다.")
    if body.direction not in {"encrypt", "decrypt"}:
        raise HTTPException(status_code=400, detail="direction은 encrypt 또는 decrypt여야 합니다.")

    count = 0

    if body.category == "faq":
        rows = db.query(FaqRecord).filter(FaqRecord.is_active.is_(True)).all()
        fields = ["category", "question", "answer", "keywords_json", "aliases_json", "search_hints_json", "source_files_json"]
        for row in rows:
            changed = False
            for field in fields:
                value = getattr(row, field)
                if not value:
                    continue
                if body.direction == "decrypt" and value.startswith(ENCRYPTED_PREFIX):
                    decrypted = decrypt_if_needed(value)
                    if decrypted:  # 복호화 실패(키 불일치 등) 시 ""로 덮어쓰지 않고 원본 유지 → 데이터 손실 방지
                        setattr(row, field, decrypted)
                        changed = True
                elif body.direction == "encrypt" and not value.startswith(ENCRYPTED_PREFIX):
                    setattr(row, field, encrypt(value))
                    changed = True
            if changed:
                count += 1
        db.commit()

    elif body.category == "prompt":
        rows = db.query(PromptConfig).all()
        for row in rows:
            value = row.content
            if not value:
                continue
            if body.direction == "decrypt" and value.startswith(ENCRYPTED_PREFIX):
                row.content = decrypt_if_needed(value) or value
                count += 1
            elif body.direction == "encrypt" and not value.startswith(ENCRYPTED_PREFIX):
                row.content = encrypt(value)
                count += 1
        db.commit()

    elif body.category == "document":
        rows = db.query(DocumentRecord).filter(DocumentRecord.is_deleted.is_(False)).all()
        for row in rows:
            changed = False
            for field in ["original_filename", "review_note", "error_message"]:
                value = getattr(row, field)
                if not value:
                    continue
                if body.direction == "decrypt" and value.startswith(ENCRYPTED_PREFIX):
                    decrypted = decrypt_if_needed(value)
                    if decrypted:  # 복호화 실패(키 불일치 등) 시 ""로 덮어쓰지 않고 원본 유지 → 데이터 손실 방지
                        setattr(row, field, decrypted)
                        changed = True
                elif body.direction == "encrypt" and not value.startswith(ENCRYPTED_PREFIX):
                    setattr(row, field, encrypt(value))
                    changed = True
            if changed:
                count += 1
        db.commit()

    action = "암호화" if body.direction == "encrypt" else "복호화"
    create_audit_log(db, f"encryption_migrated_{body.direction}", "system", body.category, f"{count}개 처리")
    return {"message": f"{count}개 레코드를 {action}했습니다.", "count": count, "category": body.category, "direction": body.direction}
