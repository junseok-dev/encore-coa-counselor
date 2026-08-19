import io
import unittest

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from app.db.database import Base
from app.db.models import ChatMessage, ChatSession
from app.routers import admin


class SessionExportTest(unittest.TestCase):
    def setUp(self):
        self.engine = create_engine(
            "sqlite://",
            connect_args={"check_same_thread": False},
            poolclass=StaticPool,
        )
        Base.metadata.create_all(self.engine)
        self.session_factory = sessionmaker(bind=self.engine)
        self.db = self.session_factory()

    def tearDown(self):
        self.db.close()
        self.engine.dispose()

    def test_exports_the_complete_session_conversation(self):
        self.db.add(ChatSession(id="session-1", encrypted_user_name="홍길동", message_count=2))
        self.db.add_all([
            ChatMessage(session_id="session-1", role="user", content="수강 신청 방법을 알려주세요.", source="user"),
            ChatMessage(session_id="session-1", role="assistant", content="신청 페이지에서 접수할 수 있습니다.", source="faq"),
        ])
        self.db.commit()

        response = admin.export_session("session-1", self.db, None)
        workbook = load_workbook(io.BytesIO(response.body), read_only=True, data_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
        workbook.close()

        self.assertEqual("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", response.media_type)
        self.assertIn("chat_session_session-1_", response.headers["content-disposition"])
        self.assertEqual(("세션 ID", "session-1", None, None, None), rows[0])
        self.assertEqual(("순서", "역할", "내용", "응답 출처", "작성 시각"), rows[5])
        self.assertEqual("수강 신청 방법을 알려주세요.", rows[6][2])
        self.assertEqual("신청 페이지에서 접수할 수 있습니다.", rows[7][2])

    def test_missing_session_returns_not_found(self):
        with self.assertRaises(HTTPException) as context:
            admin.export_session("missing", self.db, None)

        self.assertEqual(404, context.exception.status_code)


if __name__ == "__main__":
    unittest.main()
